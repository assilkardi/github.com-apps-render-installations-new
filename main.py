import asyncio
import html
import json
import logging
import os
import re
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import pandas as pd
import requests
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.constants import ChatAction, ParseMode
from telegram.error import BadRequest
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

from config import (
    TOKEN,
    SERPAPI_KEY,
    PAPPERS_API_KEY,
    INSEE_CLIENT_ID,
    INSEE_CLIENT_SECRET,
    OWNER_USER_ID,
    ADMIN_USER_IDS,
    SERPER_API_KEY,
)
# --------------------------------------------------
# BRANDING / UI
# --------------------------------------------------
BOT_BRAND_NAME = "LeadGen Premium"
BOT_BRAND_TAGLINE = "Prospection intelligente • Recherche ciblée • Export propre"
BOT_BRAND_ACCENT = "✨"
BOT_BRAND_SUPPORT = "Interface premium Telegram"

BRAND_IMAGE = os.getenv("BRAND_IMAGE", "").strip()

# --------------------------------------------------
# CONFIG
# --------------------------------------------------
MAX_RESULTS = 200
DISPLAY_PAGE_SIZE = 20
COMPANY_PAGE_SIZE = 5
SERP_BATCH_SIZE = 10
MAX_MESSAGE_SAFE = 3800
MAX_RESULTS_OPTIONS = [20, 50, 100, 200]

EXPORT_DIR = "exports"
CACHE_FILE = "search_cache.json"
ACCESS_CONTROL_FILE = "access_control.json"
REQUEST_TIMEOUT = 20

MAX_COMPANY_STRONG_RESULTS = 5
MAX_COMPANY_TOTAL_RESULTS = 30

CACHE_TTL_DEFAULT = 1800
CACHE_TTL_COMPANY = 86400
CACHE_TTL_PERSON = 43200
CACHE_TTL_PROSPECT = 43200
CACHE_TTL_WEB = 86400
CACHE_TTL_ANNUAIRE = 86400
CACHE_TTL_NEGATIVE_WEB = 900
CACHE_TTL_NEGATIVE_SEARCH = 1200

REQUEST_CONNECT_TIMEOUT = 6
REQUEST_READ_TIMEOUT = 18
REQUEST_TIMEOUT_TUPLE = (REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT)
SERPER_MAX_RETRIES = 2
SERPER_RETRY_DELAYS = [0.8, 1.6]
SERPAPI_COOLDOWN_SECONDS = 300
PROVIDER_FAIL_COOLDOWN_SECONDS = 180
PROVIDER_MAX_CONSECUTIVE_FAILURES = 3
MAX_PROSPECT_QUERY_VARIANTS = 4
MAX_COMPANY_QUERY_VARIANTS = 4
COMPANY_WEB_MAX_WORKERS = 2
COMPANY_WEB_MAX_ENRICH_CANDIDATES = 8
COMPANY_WEB_QUERY_BUDGET = 4
ACCESS_HISTORY_MAX_ITEMS = 500

SEARCH_CACHE: Dict[str, Dict[str, object]] = {}
PROVIDER_STATS = {
    "cache": 0,
    "cache_negative": 0,
    "annuaire_api": 0,
    "serper": 0,
    "serpapi": 0,
    "serpapi_429": 0,
    "serpapi_cooldown_hits": 0,
    "provider_skips": 0,
    "company_query_budget_hits": 0,
}

PROVIDER_STATE: Dict[str, Dict[str, Any]] = {
    "serper": {"cooldown_until": 0, "failures": 0, "last_error": ""},
    "serpapi": {"cooldown_until": 0, "failures": 0, "last_error": ""},
}

ACCESS_STATE: Dict[str, Dict[str, Dict[str, Any]]] = {
    "approved_users": {},
    "pending_users": {},
    "blacklist": {},
    "history": [],
}

ANNUAIRE_API_BASE = "https://recherche-entreprises.api.gouv.fr"
ANNUAIRE_USER_AGENT = "TelegramProspectBot/2.0"
def provider_available(provider: str) -> bool:
    """
    Vérifie si un provider est disponible (pas en cooldown)
    """
    state = PROVIDER_STATE.get(provider)
    if not state:
        return True

    cooldown_until = state.get("cooldown_until")
    if cooldown_until and time.time() < cooldown_until:
        return False

    return True
def mark_provider_failure(provider: str, is_429: bool = False):
    """
    Marque un provider comme en échec et applique cooldown si nécessaire
    """
    state = PROVIDER_STATE.setdefault(provider, {
        "failures": 0,
        "cooldown_until": 0
    })

    state["failures"] += 1

    # Si 429 → cooldown immédiat (5 min)
    if is_429:
        state["cooldown_until"] = time.time() + 300  # 5 minutes
    else:
        # petit circuit breaker
        if state["failures"] >= 3:
            state["cooldown_until"] = time.time() + 120  # 2 minutes
            state["failures"] = 0

def mark_provider_success(provider: str):
    state = PROVIDER_STATE.get(provider)
    if state:
        state["failures"] = 0
        state["cooldown_until"] = 0

# --------------------------------------------------
# LOGGING
# --------------------------------------------------
logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.WARNING,
)

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("telegram").setLevel(logging.WARNING)
logging.getLogger("telegram.ext").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)

# --------------------------------------------------
# CACHE DISQUE
# --------------------------------------------------
def load_cache() -> None:
    global SEARCH_CACHE
    if not os.path.exists(CACHE_FILE):
        SEARCH_CACHE = {}
        return

    try:
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            SEARCH_CACHE = json.load(f)
    except Exception as e:
        logger.warning("Impossible de charger le cache disque: %s", e)
        SEARCH_CACHE = {}


def save_cache() -> None:
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(SEARCH_CACHE, f, ensure_ascii=False)
    except Exception as e:
        logger.warning("Impossible de sauvegarder le cache disque: %s", e)


def load_access_state() -> None:
    global ACCESS_STATE
    if not os.path.exists(ACCESS_CONTROL_FILE):
        ACCESS_STATE = {"approved_users": {}, "pending_users": {}, "blacklist": {}, "history": []}
        return

    try:
        with open(ACCESS_CONTROL_FILE, "r", encoding="utf-8") as f:
            payload = json.load(f)
            if isinstance(payload, dict):
                ACCESS_STATE = {
                    "approved_users": payload.get("approved_users", {}) or {},
                    "pending_users": payload.get("pending_users", {}) or {},
                    "blacklist": payload.get("blacklist", {}) or {},
                    "history": payload.get("history", []) or [],
                }
            else:
                ACCESS_STATE = {"approved_users": {}, "pending_users": {}, "blacklist": {}, "history": []}
    except Exception as e:
        logger.warning("Impossible de charger les accès: %s", e)
        ACCESS_STATE = {"approved_users": {}, "pending_users": {}, "blacklist": {}, "history": []}

    cleanup_access_state(save=False)


def save_access_state() -> None:
    try:
        with open(ACCESS_CONTROL_FILE, "w", encoding="utf-8") as f:
            json.dump(ACCESS_STATE, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning("Impossible de sauvegarder les accès: %s", e)


def now_ts() -> int:
    return int(time.time())


def user_key(user_id: int) -> str:
    return str(int(user_id))


def user_display_name(first_name: str = "", username: str = "") -> str:
    if first_name and username:
        return f"{first_name} (@{username})"
    if username:
        return f"@{username}"
    return first_name or "Utilisateur"


def build_user_info(user) -> Dict[str, Any]:
    return {
        "user_id": int(user.id),
        "first_name": getattr(user, "first_name", "") or "",
        "last_name": getattr(user, "last_name", "") or "",
        "username": getattr(user, "username", "") or "",
    }


def build_user_profile_link(user_info: Dict[str, Any]) -> str:
    first_name = html.escape(user_info.get("first_name", "") or "Utilisateur")
    username = (user_info.get("username") or "").strip()
    user_id = int(user_info.get("user_id"))
    if username:
        return f'<a href="https://t.me/{html.escape(username)}">{first_name}</a> (@{html.escape(username)})'
    return f'<a href="tg://user?id={user_id}">{first_name}</a>'


def access_request_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔓 Demander l'accès", callback_data="access_request")],
    ])


def profile_action_rows(user_info: Dict[str, Any]) -> List[List[InlineKeyboardButton]]:
    rows: List[List[InlineKeyboardButton]] = []
    username = (user_info.get("username") or "").strip()
    if username:
        rows.append([InlineKeyboardButton("👤 Ouvrir le profil", url=f"https://t.me/{username}")])
    return rows


def access_manage_keyboard(target_user_id: int, pending: bool = True, blacklisted: bool = False) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("✅ Accès illimité", callback_data=f"access:grant:perm:{target_user_id}")],
        [
            InlineKeyboardButton("5 min", callback_data=f"access:grant:m5:{target_user_id}"),
            InlineKeyboardButton("10 min", callback_data=f"access:grant:m10:{target_user_id}"),
            InlineKeyboardButton("15 min", callback_data=f"access:grant:m15:{target_user_id}"),
        ],
        [
            InlineKeyboardButton("5 recherches", callback_data=f"access:grant:s5:{target_user_id}"),
            InlineKeyboardButton("10 recherches", callback_data=f"access:grant:s10:{target_user_id}"),
            InlineKeyboardButton("25 recherches", callback_data=f"access:grant:s25:{target_user_id}"),
        ],
    ]
    if blacklisted:
        rows.append([InlineKeyboardButton("♻️ Retirer blacklist", callback_data=f"access:unblacklist:{target_user_id}")])
    elif pending:
        rows.append([
            InlineKeyboardButton("❌ Refuser", callback_data=f"access:deny:{target_user_id}"),
            InlineKeyboardButton("🚫 Blacklist", callback_data=f"access:blacklist:{target_user_id}"),
        ])
    else:
        rows.append([
            InlineKeyboardButton("🗑️ Supprimer l'accès", callback_data=f"access:revoke:{target_user_id}"),
            InlineKeyboardButton("🚫 Blacklist", callback_data=f"access:blacklist:{target_user_id}"),
        ])
    return InlineKeyboardMarkup(rows)


def format_access_badge(record: Optional[Dict[str, Any]]) -> str:
    if not record:
        return "Aucun accès"
    expires_at = record.get("expires_at")
    remaining_searches = record.get("remaining_searches")
    if expires_at:
        dt = datetime.fromtimestamp(int(expires_at))
        return f"Jusqu'au {dt.strftime('%d/%m %H:%M')}"
    if isinstance(remaining_searches, int):
        return f"{remaining_searches} recherche(s) restante(s)"
    return "Accès illimité"


def append_access_history(event: str, user_info: Dict[str, Any], actor_id: Optional[int] = None, meta: Optional[Dict[str, Any]] = None) -> None:
    history = ACCESS_STATE.setdefault("history", [])
    entry = {
        "ts": now_ts(),
        "event": event,
        "user_id": int(user_info.get("user_id") or 0),
        "first_name": user_info.get("first_name", "") or "",
        "username": user_info.get("username", "") or "",
        "actor_id": int(actor_id) if actor_id else None,
        "meta": meta or {},
    }
    history.append(entry)
    if len(history) > ACCESS_HISTORY_MAX_ITEMS:
        del history[:-ACCESS_HISTORY_MAX_ITEMS]


def cleanup_access_state(save: bool = True) -> None:
    approved = ACCESS_STATE.setdefault("approved_users", {})
    pending = ACCESS_STATE.setdefault("pending_users", {})
    ACCESS_STATE.setdefault("blacklist", {})
    history = ACCESS_STATE.setdefault("history", [])
    changed = False
    current = now_ts()

    for uid, record in list(approved.items()):
        expires_at = record.get("expires_at")
        remaining_searches = record.get("remaining_searches")
        if expires_at and int(expires_at) <= current:
            append_access_history("expired", record, meta={"reason": "time"})
            approved.pop(uid, None)
            changed = True
            continue
        if isinstance(remaining_searches, int) and remaining_searches <= 0:
            append_access_history("expired", record, meta={"reason": "credits"})
            approved.pop(uid, None)
            changed = True

    for uid, record in list(pending.items()):
        request_at = int(record.get("request_at") or current)
        if current - request_at > 7 * 86400:
            pending.pop(uid, None)
            append_access_history("pending_purged", record)
            changed = True

    if len(history) > ACCESS_HISTORY_MAX_ITEMS:
        del history[:-ACCESS_HISTORY_MAX_ITEMS]
        changed = True

    if changed and save:
        save_access_state()


def get_approved_record(user_id: int) -> Optional[Dict[str, Any]]:
    cleanup_access_state(save=True)
    return ACCESS_STATE.get("approved_users", {}).get(user_key(user_id))


def get_pending_record(user_id: int) -> Optional[Dict[str, Any]]:
    cleanup_access_state(save=True)
    return ACCESS_STATE.get("pending_users", {}).get(user_key(user_id))


def get_blacklist_record(user_id: int) -> Optional[Dict[str, Any]]:
    cleanup_access_state(save=True)
    return ACCESS_STATE.get("blacklist", {}).get(user_key(user_id))


def is_blacklisted(user_id: int) -> bool:
    return get_blacklist_record(user_id) is not None


def has_runtime_access(user_id: int) -> bool:
    if is_admin(user_id):
        return True
    if is_blacklisted(user_id):
        return False
    return get_approved_record(user_id) is not None


def grant_user_access(user_info: Dict[str, Any], granted_by: int, mode: str) -> Dict[str, Any]:
    approved = ACCESS_STATE.setdefault("approved_users", {})
    pending = ACCESS_STATE.setdefault("pending_users", {})

    uid = user_key(user_info["user_id"])
    record = {
        "user_id": int(user_info["user_id"]),
        "first_name": user_info.get("first_name", ""),
        "last_name": user_info.get("last_name", ""),
        "username": user_info.get("username", ""),
        "granted_at": now_ts(),
        "granted_by": int(granted_by),
        "expires_at": None,
        "remaining_searches": None,
        "mode": mode,
    }

    if mode.startswith("m") and mode[1:].isdigit():
        minutes = int(mode[1:])
        record["expires_at"] = now_ts() + minutes * 60
    elif mode.startswith("s") and mode[1:].isdigit():
        record["remaining_searches"] = int(mode[1:])

    approved[uid] = record
    pending.pop(uid, None)
    ACCESS_STATE.setdefault("blacklist", {}).pop(uid, None)
    append_access_history("granted", record, actor_id=granted_by, meta={"mode": mode})
    save_access_state()
    return record


def deny_user_request(user_id: int, denied_by: Optional[int] = None) -> None:
    record = ACCESS_STATE.setdefault("pending_users", {}).pop(user_key(user_id), None)
    if record:
        append_access_history("denied", record, actor_id=denied_by)
    save_access_state()


def revoke_user_access(user_id: int, revoked_by: Optional[int] = None) -> None:
    key = user_key(user_id)
    record = ACCESS_STATE.setdefault("approved_users", {}).pop(key, None) or ACCESS_STATE.setdefault("pending_users", {}).pop(key, None)
    if record:
        append_access_history("revoked", record, actor_id=revoked_by)
    save_access_state()


def blacklist_user_access(user_info: Dict[str, Any], blacklisted_by: int, reason: str = "") -> Dict[str, Any]:
    key = user_key(user_info["user_id"])
    ACCESS_STATE.setdefault("approved_users", {}).pop(key, None)
    ACCESS_STATE.setdefault("pending_users", {}).pop(key, None)
    record = {
        "user_id": int(user_info["user_id"]),
        "first_name": user_info.get("first_name", ""),
        "last_name": user_info.get("last_name", ""),
        "username": user_info.get("username", ""),
        "blacklisted_at": now_ts(),
        "blacklisted_by": int(blacklisted_by),
        "reason": clean_spaces(reason),
    }
    ACCESS_STATE.setdefault("blacklist", {})[key] = record
    append_access_history("blacklisted", record, actor_id=blacklisted_by, meta={"reason": clean_spaces(reason)})
    save_access_state()
    return record


def unblacklist_user_access(user_id: int, actor_id: Optional[int] = None) -> Optional[Dict[str, Any]]:
    record = ACCESS_STATE.setdefault("blacklist", {}).pop(user_key(user_id), None)
    if record:
        append_access_history("unblacklisted", record, actor_id=actor_id)
        save_access_state()
    return record


def register_pending_request(user_info: Dict[str, Any]) -> bool:
    if is_blacklisted(int(user_info["user_id"])):
        return False
    pending = ACCESS_STATE.setdefault("pending_users", {})
    key = user_key(user_info["user_id"])
    if key in pending:
        return False

    pending[key] = {
        **user_info,
        "request_at": now_ts(),
    }
    append_access_history("requested", user_info)
    save_access_state()
    return True


async def notify_user_access_granted(bot, user_id: int, record: Dict[str, Any]) -> None:
    try:
        await bot.send_message(
            chat_id=user_id,
            text=(
                "✅ <b>Accès accordé</b>\n\n"
                f"Statut : {format_access_badge(record)}\n\n"
                "Tu peux maintenant utiliser le bot avec /start."
            ),
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        logger.warning("Impossible de notifier l'utilisateur %s: %s", user_id, e)


async def notify_user_access_denied(bot, user_id: int) -> None:
    try:
        await bot.send_message(
            chat_id=user_id,
            text="❌ Ta demande d'accès a été refusée.",
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        logger.warning("Impossible de notifier le refus à %s: %s", user_id, e)


async def notify_user_access_revoked(bot, user_id: int) -> None:
    try:
        await bot.send_message(
            chat_id=user_id,
            text="⛔ Ton accès au bot a été retiré.",
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        logger.warning("Impossible de notifier la révocation à %s: %s", user_id, e)


async def send_access_request_to_owner(bot, user_info: Dict[str, Any]) -> bool:
    if not OWNER_USER_ID:
        return False

    profile_link = build_user_profile_link(user_info)
    text = (
        "🔐 <b>Nouvelle demande d'accès</b>\n\n"
        f"👤 Profil : {profile_link}\n"
        f"🆔 ID : <code>{user_info['user_id']}</code>\n"
        f"📝 Nom : {esc(user_display_name(user_info.get('first_name', ''), user_info.get('username', '')))}\n"
    )
    rows = []
    rows.extend(profile_action_rows(user_info))
    rows.extend(access_manage_keyboard(int(user_info["user_id"]), pending=True).inline_keyboard)
    try:
        await bot.send_message(
            chat_id=OWNER_USER_ID,
            text=text,
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(rows),
            disable_web_page_preview=True,
        )
        return True
    except Exception as e:
        logger.warning("Impossible d'envoyer la demande à l'owner: %s", e)
        return False


async def deny_access(update: Update, context: Optional[ContextTypes.DEFAULT_TYPE] = None) -> None:
    message = update.effective_message
    user = update.effective_user
    if not message or not user:
        return

    if is_blacklisted(user.id):
        text = (
            "⛔ <b>Accès bloqué</b>\n\n"
            "Ton accès à ce bot est actuellement bloqué."
        )
        markup = None
    else:
        pending = get_pending_record(user.id)
        if pending:
            text = (
                "⏳ <b>Demande en attente</b>\n\n"
                "Ta demande d'accès a déjà été envoyée. J'attends la réponse de l'administrateur."
            )
            markup = None
        else:
            text = (
                "🔒 <b>Accès privé</b>\n\n"
                "Tu n'as pas encore accès à ce bot.\n"
                "Appuie sur le bouton ci-dessous pour m'envoyer une demande d'accès."
            )
            markup = access_request_keyboard()

    await message.reply_text(text, parse_mode=ParseMode.HTML, reply_markup=markup)


async def require_access(update: Update, context: Optional[ContextTypes.DEFAULT_TYPE] = None) -> bool:
    user = update.effective_user
    if not user:
        return False
    if has_runtime_access(user.id):
        return True
    await deny_access(update, context)
    return False


def consume_search_credit(user_id: int) -> Optional[Dict[str, Any]]:
    if is_admin(user_id):
        return None

    key = user_key(user_id)
    record = ACCESS_STATE.get("approved_users", {}).get(key)
    if not record:
        return None

    remaining_searches = record.get("remaining_searches")
    if isinstance(remaining_searches, int):
        remaining_searches -= 1
        record["remaining_searches"] = remaining_searches
        if remaining_searches <= 0:
            ACCESS_STATE["approved_users"].pop(key, None)
            save_access_state()
            return {**record, "remaining_searches": 0, "access_revoked": True}
        save_access_state()
        return record

    expires_at = record.get("expires_at")
    if expires_at and int(expires_at) <= now_ts():
        ACCESS_STATE["approved_users"].pop(key, None)
        save_access_state()
        return {**record, "access_revoked": True}

    return record


async def after_search_usage(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if not user:
        return

    record = consume_search_credit(user.id)
    if not record:
        return

    message = update.effective_message
    if not message:
        return

    if record.get("access_revoked"):
        await message.reply_text(
            "⛔ C'était ta dernière recherche autorisée. Ton accès a expiré.",
            parse_mode=ParseMode.HTML,
        )
        return

    remaining_searches = record.get("remaining_searches")
    if isinstance(remaining_searches, int):
        await message.reply_text(
            f"ℹ️ Il te reste <b>{remaining_searches}</b> recherche(s).",
            parse_mode=ParseMode.HTML,
        )


def get_cache_ttl(cache_key: str, cached: Optional[Dict[str, object]] = None) -> int:
    is_negative = False
    if cached:
        data = cached.get("data") if isinstance(cached, dict) else None
        if isinstance(data, dict):
            is_negative = bool(data.get("negative_cache"))

    if cache_key.startswith("company__"):
        return CACHE_TTL_NEGATIVE_SEARCH if is_negative else CACHE_TTL_COMPANY
    if cache_key.startswith("annuaire__"):
        return CACHE_TTL_NEGATIVE_SEARCH if is_negative else CACHE_TTL_ANNUAIRE
    if cache_key.startswith("person__"):
        return CACHE_TTL_NEGATIVE_SEARCH if is_negative else CACHE_TTL_PERSON
    if cache_key.startswith("prospect__"):
        return CACHE_TTL_NEGATIVE_SEARCH if is_negative else CACHE_TTL_PROSPECT
    if cache_key.startswith("web__"):
        return CACHE_TTL_NEGATIVE_WEB if is_negative else CACHE_TTL_WEB
    return CACHE_TTL_DEFAULT


def get_cache_payload(cache_key: str) -> Optional[Dict[str, object]]:
    cached = SEARCH_CACHE.get(cache_key)
    if not cached:
        return None

    now = time.time()
    ts = float(cached.get("ts", 0))
    ttl = get_cache_ttl(cache_key, cached)

    if now - ts < ttl:
        data = cached.get("data")
        if isinstance(data, dict) and data.get("negative_cache"):
            PROVIDER_STATS["cache_negative"] += 1
        else:
            PROVIDER_STATS["cache"] += 1
        return data

    return None


def set_cache_payload(cache_key: str, data: Dict[str, object]) -> None:
    SEARCH_CACHE[cache_key] = {
        "ts": time.time(),
        "data": data,
    }
    save_cache()

# --------------------------------------------------
# UI HELPERS
# --------------------------------------------------
def esc(text: Any) -> str:
    return html.escape(str(text or ""))


def brand_header_text() -> str:
    return (
        f"<b>{esc(BOT_BRAND_NAME)}</b>\n"
        f"{esc(BOT_BRAND_TAGLINE)}\n\n"
        f"{BOT_BRAND_ACCENT} <i>{esc(BOT_BRAND_SUPPORT)}</i>"
    )

def main_menu_text() -> str:
    return (
        f"✨ <b>{esc(BOT_BRAND_NAME)}</b>\n"
        f"<i>{esc(BOT_BRAND_TAGLINE)}</i>\n\n"
        "━━━━━━━━━━━━━━━━━━━\n\n"
        "🚀 <b>Prospects LinkedIn</b>\n"
        "<i>Identifie des profils qualifiés à partir de mots-clés avancés.</i>\n\n"
        "👤 <b>Recherche personne</b>\n"
        "<i>Retrouve un profil précis rapidement.</i>\n\n"
        "🏢 <b>Entreprise & dirigeant</b>\n"
        "<i>Accède aux données clés et aux décideurs.</i>\n\n"
        "━━━━━━━━━━━━━━━━━━━\n\n"
        "⚡ <b>Fonctionnalités</b>\n"
        "• Recherche rapide ou avancée\n"
        "• Filtres intelligents (pays, niveau, métier)\n"
        "• Résultats enrichis et scorés\n"
        "• Export Excel professionnel\n\n"
        "👇 <b>Sélectionne une option ci-dessous</b>"
    )


def prospects_intro_text() -> str:
    return (
        "<b>Module Prospects</b>\n\n"
        "Choisis ton mode :\n"
        "• <b>Rapide</b> : tu envoies juste ton besoin.\n"
        "• <b>Avancé</b> : tu ajoutes des filtres, une zone géographique et plus d’options."
    )


def prospect_query_help_text(mode_variant: str = "quick") -> str:
    examples = (
        "<b>Exemples :</b>\n"
        "• <code>développeur java</code>\n"
        "• <code>business developer saas</code>\n"
        "• <code>mission expérience 5 ans développeur react</code>"
    )

    if mode_variant == "advanced":
        return (
            "🎯 <b>Recherche prospects avancée</b>\n\n"
            "Envoie maintenant ton besoin ou ta description de poste.\n\n"
            f"{examples}\n\n"
            "👉 Ensuite, je te proposerai les filtres détaillés."
        )

    return (
        "⚡ <b>Recherche prospects rapide</b>\n\n"
        "Envoie directement ton besoin.\n\n"
        f"{examples}\n\n"
        "👉 Je te laisserai ensuite choisir une zone géographique et l’export."
    )


def person_intro_text() -> str:
    return (
        "<b>Recherche personne</b>\n\n"
        "Choisis le type de recherche à effectuer :"
    )


def filters_help_text() -> str:
    return (
        "<b>Filtres avancés</b>\n\n"
        "Format attendu :\n"
        "<code>ville=Paris,pays=France,entreprise=Google,poste=Sales,seniorite=manager</code>\n\n"
        "Tu peux aussi répondre simplement : <b>aucun</b>."
    )


def person_filters_help_text() -> str:
    return (
        "<b>Filtres optionnels</b>\n\n"
        "Format attendu :\n"
        "<code>ville=Paris,pays=France,entreprise=Airbus</code>\n\n"
        "Tu peux aussi répondre simplement : <b>aucun</b>."
    )


def excel_choice_text() -> str:
    return (
        "<b>Export Excel</b>\n\n"
        "Souhaites-tu générer également un <b>fichier Excel complet</b> ?"
    )


def loading_search_text(search_mode: str) -> str:
    if search_mode == "person":
        return (
            "🔎 <b>Recherche LinkedIn en cours…</b>\n"
            "Merci de patienter quelques secondes."
        )
    return (
        "🔎 <b>Recherche prospects en cours…</b>\n"
        "Préparation des meilleurs profils."
    )


def company_search_loading_text() -> str:
    return (
        "🏢 <b>Recherche entreprise en cours…</b>\n"
        "Analyse des sources officielles et enrichissement web."
    )


def no_result_text() -> str:
    return (
        "Aucun résultat exploitable trouvé.\n"
        "Essaie une autre orthographe, un autre mot-clé ou des filtres plus larges."
    )


def search_done_text(count: int, mode: str) -> str:
    if mode == "person":
        return f"✅ <b>{count}</b> profil(s) trouvé(s)."
    if mode == "company":
        return f"✅ <b>{count}</b> résultat(s) entreprise trouvé(s), classé(s) par pertinence."
    return f"✅ <b>{count}</b> profil(s) trouvé(s)."


def search_summary_text(mode: str, query: str, filters: Dict[str, str], count: int) -> str:
    filters_txt = ", ".join(f"{k}={v}" for k, v in filters.items()) if filters else "aucun"
    label = {
        "prospect": "Prospects LinkedIn",
        "person": "Recherche personne",
        "company": "Entreprise / dirigeant",
    }.get(mode, mode)

    return (
        "<b>Recherche terminée</b>\n\n"
        f"• Module : <b>{esc(label)}</b>\n"
        f"• Requête : <code>{esc(query)}</code>\n"
        f"• Filtres : <code>{esc(filters_txt)}</code>\n"
        f"• Résultats : <b>{count}</b>"
    )

# --------------------------------------------------
# CLAVIERS
# --------------------------------------------------
def menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🚀 Prospects LinkedIn", callback_data="menu_prospects")],
        [InlineKeyboardButton("👤 Recherche personne", callback_data="menu_person_menu")],
        [InlineKeyboardButton("🏢 Entreprise / dirigeant", callback_data="menu_company_direct")],
        [
            InlineKeyboardButton("🧠 Aide", callback_data="menu_help"),
            InlineKeyboardButton("📊 Dernière recherche", callback_data="menu_last"),
        ],
    ])


def back_to_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Retour menu", callback_data="menu_home")]])


def last_search_keyboard(can_relaunch: bool = False) -> InlineKeyboardMarkup:
    rows = []
    if can_relaunch:
        rows.append([
            InlineKeyboardButton("🔁 Relancer", callback_data="action_rerun_last"),
            InlineKeyboardButton("📊 Exporter", callback_data="action_export_last"),
        ])
    rows.append([InlineKeyboardButton("⬅️ Retour menu", callback_data="menu_home")])
    return InlineKeyboardMarkup(rows)


def prospects_mode_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⚡ Recherche rapide", callback_data="prospect_mode_quick")],
        [InlineKeyboardButton("🎯 Recherche avancée", callback_data="prospect_mode_advanced")],
        [InlineKeyboardButton("⬅️ Retour menu", callback_data="menu_home")],
    ])


def prospect_geo_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🇫🇷 France", callback_data="prospect_geo_france"),
            InlineKeyboardButton("🌍 Maghreb", callback_data="prospect_geo_maghreb"),
        ],
        [
            InlineKeyboardButton("🇪🇺 Europe", callback_data="prospect_geo_europe"),
            InlineKeyboardButton("🌐 International", callback_data="prospect_geo_international"),
        ],
        [InlineKeyboardButton("➖ Aucun filtre", callback_data="prospect_geo_none")],
        [InlineKeyboardButton("⬅️ Retour menu", callback_data="menu_home")],
    ])


def max_results_keyboard() -> InlineKeyboardMarkup:
    rows = []
    current = []
    for value in MAX_RESULTS_OPTIONS:
        current.append(InlineKeyboardButton(str(value), callback_data=f"max_results:{value}"))
        if len(current) == 2:
            rows.append(current)
            current = []
    if current:
        rows.append(current)
    rows.append([InlineKeyboardButton("⬅️ Retour menu", callback_data="menu_home")])
    return InlineKeyboardMarkup(rows)


def person_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔗 Recherche LinkedIn", callback_data="person_linkedin")],
        [InlineKeyboardButton("🏢 Recherche entreprise", callback_data="person_company")],
        [InlineKeyboardButton("⬅️ Retour menu", callback_data="menu_home")],
    ])


def excel_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📊 Oui, exporter", callback_data="excel_yes"),
            InlineKeyboardButton("➖ Non", callback_data="excel_no"),
        ],
        [InlineKeyboardButton("⬅️ Retour menu", callback_data="menu_home")],
    ])


def fuzzy_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Oui", callback_data="fuzzy_yes"),
            InlineKeyboardButton("❌ Non", callback_data="fuzzy_no"),
        ],
        [InlineKeyboardButton("⬅️ Retour menu", callback_data="menu_home")],
    ])


def pagination_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Afficher 20 résultats de plus", callback_data="more")],
        [InlineKeyboardButton("📊 Export Excel", callback_data="action_export_last"), InlineKeyboardButton("🏠 Menu", callback_data="menu_home")],
    ])


def company_page_keyboard(current_page: int, total_results: int) -> InlineKeyboardMarkup:
    start_index = current_page * COMPANY_PAGE_SIZE
    end_index = min(start_index + COMPANY_PAGE_SIZE, total_results)
    rows = []
    detail_buttons = []
    for i in range(start_index, end_index):
        local_number = i - start_index + 1
        detail_buttons.append(InlineKeyboardButton(f"📄 Détails {local_number}", callback_data=f"company_detail_{i}"))
    for i in range(0, len(detail_buttons), 2):
        rows.append(detail_buttons[i:i + 2])
    nav_row = []
    if current_page > 0:
        nav_row.append(InlineKeyboardButton("⬅️ Précédents", callback_data="company_prev"))
    if end_index < total_results:
        nav_row.append(InlineKeyboardButton("Suivants ➡️", callback_data="company_next"))
    if nav_row:
        rows.append(nav_row)
    rows.append([InlineKeyboardButton("🏠 Menu", callback_data="menu_home")])
    return InlineKeyboardMarkup(rows)


def company_detail_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🗑️ Fermer", callback_data="company_back")],
        [InlineKeyboardButton("🏠 Menu", callback_data="menu_home")],
    ])

# --------------------------------------------------
# OUTILS GÉNÉRAUX
# --------------------------------------------------
def sanitize_filename(text: str) -> str:
    text = re.sub(r"[^\w\-]+", "_", text.strip(), flags=re.UNICODE)
    return text[:80] or "recherche"


def normalize_linkedin_url(url: str) -> str:
    if not url:
        return ""
    return url.split("?")[0].strip().rstrip("/")


def split_long_message(text: str, max_len: int = MAX_MESSAGE_SAFE) -> List[str]:
    if len(text) <= max_len:
        return [text]

    chunks = []
    current = ""

    for block in text.split("\n\n"):
        if len(current) + len(block) + 2 <= max_len:
            current = f"{current}\n\n{block}".strip()
        else:
            if current:
                chunks.append(current)
            if len(block) <= max_len:
                current = block
            else:
                for i in range(0, len(block), max_len):
                    chunks.append(block[i:i + max_len])
                current = ""

    if current:
        chunks.append(current)

    return chunks


def parse_filters(text: str) -> Dict[str, str]:
    text = (text or "").strip()
    if not text or text.lower() == "aucun":
        return {}

    result = {}
    for part in text.split(","):
        if "=" in part:
            k, v = part.split("=", 1)
            k = k.strip().lower()
            v = v.strip()
            if k and v:
                result[k] = v
    return result


def apply_prospect_geo_filter(filters: Dict[str, str], preset: str) -> Dict[str, str]:
    updated = dict(filters or {})
    updated.pop("region_scope", None)
    if preset and preset != "none":
        updated["region_scope"] = preset
    return updated


def prospect_geo_label(preset: str) -> str:
    return {
        "france": "France",
        "maghreb": "Maghreb",
        "europe": "Europe",
        "international": "International",
        "none": "Aucun filtre géographique",
    }.get(preset, "Aucun filtre géographique")


def strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def normalize_text(text: str) -> str:
    text = strip_accents(text or "")
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def smart_extract_search_query(job_text: str) -> Dict[str, Any]:
    text = normalize_text(job_text)

    # 🔥 limiter taille
    words = text.split()
    text = " ".join(words[:50])

    job = ""
    keywords = []
    city = ""

    # 🎯 JOB DETECTION
    if any(x in text for x in ["fullstack", "full stack"]):
        job = "developpeur fullstack"
    elif any(x in text for x in ["backend"]):
        job = "developpeur backend"
    elif any(x in text for x in ["frontend"]):
        job = "developpeur frontend"
    elif any(x in text for x in ["data"]):
        job = "data analyst"
    elif any(x in text for x in ["devops"]):
        job = "devops engineer"
    elif any(x in text for x in ["sales", "commercial"]):
        job = "sales manager"

    # 🎯 TECH DETECTION
    tech_map = ["python", "java", "react", "node", "aws", "azure", "kubernetes"]
    for tech in tech_map:
        if tech in text:
            keywords.append(tech)

    # 🎯 VILLE
    cities = ["paris", "lyon", "lille", "marseille", "toulouse", "nantes"]
    for c in cities:
        if c in text:
            city = c
            break

    return {
        "job": job or "developpeur",
        "keywords": keywords[:3],
        "city": city
    }


def normalize_keyword(keyword: str) -> str:
    keyword = keyword.strip()
    replacements = {
        "java script": "JavaScript",
        "javascript": "JavaScript",
        "js developer": "JavaScript developer",
        "js engineer": "JavaScript engineer",
        "rh": "RH",
        "hr": "HR",
    }
    lowered = keyword.lower()
    for bad, good in replacements.items():
        if bad in lowered:
            keyword = re.sub(bad, good, keyword, flags=re.IGNORECASE)
    return keyword.strip()


async def safe_edit_message_text(
    message_obj,
    text: str,
    reply_markup=None,
    disable_web_page_preview: bool = True,
    parse_mode: Optional[str] = ParseMode.HTML,
) -> None:
    try:
        await message_obj.edit_text(
            text=text,
            reply_markup=reply_markup,
            disable_web_page_preview=disable_web_page_preview,
            parse_mode=parse_mode,
        )
    except BadRequest as e:
        if "Message is not modified" in str(e):
            return
        raise


STOPWORDS_CACHE = {"a", "de", "des", "du", "la", "le", "les", "and", "or", "the", "for", "en", "sur", "avec", "dans", "par"}


def normalize_cache_phrase(text: str) -> str:
    normalized = normalize_text(text)
    tokens = [t for t in normalized.split() if t and t not in STOPWORDS_CACHE]
    return " ".join(tokens)


def make_cache_key(
    search_mode: str,
    base_value: str,
    custom_filters: Dict[str, str],
    start: int,
    page_size: int,
    fuzzy: bool = False,
) -> str:
    norm_base = normalize_cache_phrase(base_value)
    filters_str = "|".join(
        f"{normalize_cache_phrase(str(k))}={normalize_cache_phrase(str(v))}"
        for k, v in sorted(custom_filters.items())
        if clean_spaces(str(v))
    )
    return f"{search_mode}__{norm_base}__{filters_str}__{start}__{page_size}__{fuzzy}"


def make_web_cache_key(query: str, start: int, num: int) -> str:
    return f"web__{normalize_cache_phrase(query)}__{start}__{num}"


def parse_company_query_input(text: str) -> Tuple[str, Dict[str, str]]:
    text = (text or "").strip()
    if not text:
        return "", {}

    if "," not in text:
        return text, {}

    first_part, rest = text.split(",", 1)
    name = first_part.strip()
    filters_dict = parse_filters(rest)
    return name, filters_dict


def clean_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def first_nonempty(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            cleaned = clean_spaces(value)
            if cleaned:
                return cleaned
        else:
            as_str = clean_spaces(str(value))
            if as_str:
                return as_str
    return ""


def safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return default

# --------------------------------------------------
# NORMALISATION WEB SEARCH
# --------------------------------------------------
def normalize_organic_results(items: List[dict]) -> List[Dict[str, str]]:
    normalized = []

    for item in items or []:
        title = (item.get("title") or item.get("name") or "").strip()
        link = (
            item.get("link")
            or item.get("url")
            or item.get("href")
            or ""
        ).strip()
        snippet = (
            item.get("snippet")
            or item.get("description")
            or item.get("body")
            or ""
        ).strip()

        if not link:
            continue

        normalized.append({
            "title": title,
            "link": link,
            "snippet": snippet,
        })

    return normalized


def has_enough_results(result: Dict[str, object], minimum: int = 1) -> bool:
    organic = result.get("organic_results", []) or []
    return len(organic) >= minimum

# --------------------------------------------------
# PROVIDERS WEB SEARCH
# --------------------------------------------------
def search_with_serper(query: str, start: int = 0, num: int = 10) -> Dict[str, object]:
    if not SERPER_API_KEY:
        raise RuntimeError("SERPER_API_KEY manquante")

    page = (start // max(1, num)) + 1

    headers = {
        "X-API-KEY": SERPER_API_KEY,
        "Content-Type": "application/json",
    }

    payload = {
        "q": query,
        "gl": "fr",
        "hl": "fr",
        "num": num,
        "page": page,
    }

    last_error = None
    for attempt in range(SERPER_MAX_RETRIES + 1):
        try:
            response = requests.post(
                "https://google.serper.dev/search",
                headers=headers,
                json=payload,
                timeout=REQUEST_TIMEOUT_TUPLE,
            )
            response.raise_for_status()
            data = response.json()
            organic_results = normalize_organic_results(data.get("organic", []))
            PROVIDER_STATS["serper"] += 1
            mark_provider_success("serper")
            return {
                "organic_results": organic_results,
                "has_more": len(organic_results) >= num,
                "next_start": start + num,
                "provider": "serper",
            }
        except requests.RequestException as e:
            last_error = e
            transient = isinstance(e, (requests.Timeout, requests.ConnectionError)) or (getattr(e, "response", None) is not None and getattr(e.response, "status_code", 0) >= 500)
            if attempt < SERPER_MAX_RETRIES and transient:
                delay = min(0.3 * (attempt + 1), 1)
                time.sleep(delay)
                continue
            mark_provider_failure("serper", e)
            raise
    if last_error:
        raise last_error
    raise RuntimeError("Serper indisponible")


def search_with_serpapi(query: str, start: int = 0, num: int = 10) -> Dict[str, object]:
    if not SERPAPI_KEY:
        raise RuntimeError("SERPAPI_KEY manquante")
    if not provider_available("serpapi"):
        PROVIDER_STATS["serpapi_cooldown_hits"] += 1
        remaining = provider_cooldown_remaining("serpapi")
        raise RuntimeError(f"SerpAPI en cooldown ({remaining}s)")

    params = {
        "engine": "google",
        "q": query,
        "start": start,
        "num": num,
        "hl": "fr",
        "gl": "fr",
        "google_domain": "google.fr",
        "api_key": SERPAPI_KEY,
    }

    response = requests.get(
        "https://serpapi.com/search.json",
        params=params,
        timeout=REQUEST_TIMEOUT_TUPLE,
    )
    if response.status_code == 429:
        PROVIDER_STATS["serpapi_429"] += 1
        err = RuntimeError("SerpAPI 429 Too Many Requests")
        mark_provider_failure("serpapi", err, cooldown_seconds=SERPAPI_COOLDOWN_SECONDS)
        raise err
    response.raise_for_status()
    data = response.json()

    organic_results = normalize_organic_results(data.get("organic_results", []))
    serpapi_pagination = data.get("serpapi_pagination", {}) or {}
    next_link = serpapi_pagination.get("next")
    next_page_url = data.get("pagination", {}).get("next")
    has_more = bool(next_link or next_page_url)

    PROVIDER_STATS["serpapi"] += 1
    mark_provider_success("serpapi")

    return {
        "organic_results": organic_results,
        "has_more": has_more,
        "next_start": start + num,
        "provider": "serpapi",
    }

def search_web(query: str, start: int = 0, num: int = 10) -> Dict[str, object]:
    cache_key = make_web_cache_key(query, start, num)
    cached = get_cache_payload(cache_key)
    if cached:
        logger.info("cache_hit provider=web key=%s", cache_key)
        return cached

    result = None

    if provider_available("serper"):
        try:
            result = search_with_serper(query, start=start, num=num)
            if result.get("organic_results"):
                set_cache_payload(cache_key, result)
                logger.info("search_web provider=serper query=%s results=%s", query, len(result.get("organic_results", [])))
                return result
        except Exception as e:
            logger.warning("Serper failed: %s", e)
    else:
        logger.info("Serper skipped due to cooldown")

    if provider_available("serpapi"):
        try:
            result = search_with_serpapi(query, start=start, num=num)
            if result.get("organic_results"):
                set_cache_payload(cache_key, result)
                logger.info("search_web provider=serpapi query=%s results=%s", query, len(result.get("organic_results", [])))
                return result
        except Exception as e:
            logger.warning("SerpAPI failed: %s", e)
    else:
        PROVIDER_STATS["serpapi_cooldown_hits"] += 1
        logger.info("SerpAPI skipped due to cooldown")

    empty = {
        "organic_results": [],
        "has_more": False,
        "next_start": start + num,
        "provider": result.get("provider") if isinstance(result, dict) else "none",
        "negative_cache": True,
    }
    set_cache_payload(cache_key, empty)
    return empty
    
    
    
START_TIME = datetime.now()


def is_owner(user_id: int) -> bool:
    return bool(OWNER_USER_ID and user_id == OWNER_USER_ID)


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_USER_IDS


async def require_admin(update: Update) -> bool:
    user = update.effective_user
    if not user or not is_admin(user.id):
        message = update.effective_message
        if message:
            await message.reply_text("⛔ Action réservée à l'administrateur.", parse_mode=ParseMode.HTML)
        return False
    return True


async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_admin(update):
        return

    user = update.effective_user
    uptime = datetime.now() - START_TIME
    approved = ACCESS_STATE.get("approved_users", {})
    pending = ACCESS_STATE.get("pending_users", {})

    text = (
        "🔐 <b>Panneau admin</b>\n\n"
        f"User ID : <code>{user.id if user else 'inconnu'}</code>\n"
        f"Owner : {'oui' if user and is_owner(user.id) else 'non'}\n"
        f"Admins fixes : {len(ADMIN_USER_IDS)}\n"
        f"Accès actifs : {len(approved)}\n"
        f"Demandes en attente : {len(pending)}\n"
        f"Uptime : {str(uptime).split('.')[0]}"
    )

    if update.message:
        await update.message.reply_text(text, parse_mode=ParseMode.HTML, reply_markup=last_search_keyboard(bool(last.get("mode") and last.get("base_value"))))


async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_admin(update):
        return

    cache_size = 0
    try:
        if isinstance(SEARCH_CACHE, dict):
            cache_size = len(SEARCH_CACHE)
    except Exception:
        pass

    text = (
        "📊 <b>Stats bot</b>\n\n"
        f"Cache entries : {cache_size}\n"
        f"Accès actifs : {len(ACCESS_STATE.get('approved_users', {}))}\n"
        f"Demandes en attente : {len(ACCESS_STATE.get('pending_users', {}))}\n"
        f"Provider cache : {PROVIDER_STATS.get('cache', 0)}\n"
        f"Serper : {PROVIDER_STATS.get('serper', 0)}\n"
        f"SerpAPI : {PROVIDER_STATS.get('serpapi', 0)}"
    )

    if update.message:
        await update.message.reply_text(text, parse_mode=ParseMode.HTML, reply_markup=last_search_keyboard(bool(last.get("mode") and last.get("base_value"))))


async def whoami_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if not user or not update.message:
        return

    approved_record = get_approved_record(user.id)
    lines = [
        "👤 <b>Infos Telegram</b>",
        "",
        f"ID : <code>{user.id}</code>",
    ]

    if user.username:
        lines.append(f"Username : @{user.username}")

    if is_admin(user.id):
        lines.append("Statut : admin")
    elif is_blacklisted(user.id):
        lines.append("Accès : blacklist")
    elif approved_record:
        lines.append(f"Accès : {format_access_badge(approved_record)}")
    else:
        lines.append("Accès : non autorisé")

    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)


async def cache_clear_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_admin(update):
        return

    try:
        if isinstance(SEARCH_CACHE, dict):
            SEARCH_CACHE.clear()
            save_cache()
        await update.message.reply_text("🧹 Cache vidé.", parse_mode=ParseMode.HTML)
    except Exception as e:
        await update.message.reply_text(f"Erreur lors du vidage du cache : <code>{esc(e)}</code>", parse_mode=ParseMode.HTML)


async def users_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_admin(update):
        return

    approved = list(ACCESS_STATE.get("approved_users", {}).values())
    if not approved:
        await update.message.reply_text("Aucun utilisateur avec accès temporaire pour le moment.", parse_mode=ParseMode.HTML)
        return

    approved.sort(key=lambda x: (x.get("first_name", ""), x.get("username", "")))
    for record in approved:
        text = (
            "👥 <b>Utilisateur autorisé</b>\n\n"
            f"Nom : {esc(user_display_name(record.get('first_name', ''), record.get('username', '')))}\n"
            f"ID : <code>{record.get('user_id')}</code>\n"
            f"Statut : {format_access_badge(record)}"
        )
        await update.message.reply_text(
            text,
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(profile_action_rows(record) + access_manage_keyboard(int(record.get("user_id")), pending=False).inline_keyboard),
        )


async def pending_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_admin(update):
        return

    pending = list(ACCESS_STATE.get("pending_users", {}).values())
    if not pending:
        await update.message.reply_text("Aucune demande en attente.", parse_mode=ParseMode.HTML)
        return

    pending.sort(key=lambda x: int(x.get("request_at") or 0), reverse=True)
    for record in pending:
        text = (
            "🔐 <b>Demande en attente</b>\n\n"
            f"Nom : {esc(user_display_name(record.get('first_name', ''), record.get('username', '')))}\n"
            f"ID : <code>{record.get('user_id')}</code>"
        )
        await update.message.reply_text(
            text,
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(profile_action_rows(record) + access_manage_keyboard(int(record.get("user_id")), pending=True).inline_keyboard),
        )


async def blacklist_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_admin(update):
        return
    if not update.message:
        return
    args = context.args or []
    if not args:
        await update.message.reply_text("Usage : /blacklist <user_id>", parse_mode=ParseMode.HTML)
        return
    try:
        target_user_id = int(args[0])
    except Exception:
        await update.message.reply_text("User ID invalide.", parse_mode=ParseMode.HTML)
        return
    target_info = get_pending_record(target_user_id) or get_approved_record(target_user_id) or get_blacklist_record(target_user_id) or {"user_id": target_user_id, "first_name": "Utilisateur", "username": ""}
    blacklist_user_access(target_info, blacklisted_by=update.effective_user.id)
    await update.message.reply_text(f"🚫 Utilisateur blacklisté : <code>{target_user_id}</code>", parse_mode=ParseMode.HTML)


async def unblacklist_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_admin(update):
        return
    if not update.message:
        return
    args = context.args or []
    if not args:
        await update.message.reply_text("Usage : /unblacklist <user_id>", parse_mode=ParseMode.HTML)
        return
    try:
        target_user_id = int(args[0])
    except Exception:
        await update.message.reply_text("User ID invalide.", parse_mode=ParseMode.HTML)
        return
    removed = unblacklist_user_access(target_user_id, actor_id=update.effective_user.id)
    if not removed:
        await update.message.reply_text("Aucun utilisateur blacklisté avec cet ID.", parse_mode=ParseMode.HTML)
        return
    await update.message.reply_text(f"♻️ Utilisateur retiré de la blacklist : <code>{target_user_id}</code>", parse_mode=ParseMode.HTML)


async def blacklist_list_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_admin(update):
        return
    blocked = list(ACCESS_STATE.get("blacklist", {}).values())
    if not blocked:
        await update.message.reply_text("Aucune blacklist active.", parse_mode=ParseMode.HTML)
        return
    blocked.sort(key=lambda x: int(x.get("blacklisted_at") or 0), reverse=True)
    for record in blocked:
        reason = esc(record.get("reason", "") or "-")
        text = (
            "🚫 <b>Utilisateur blacklisté</b>\n\n"
            f"Nom : {esc(user_display_name(record.get('first_name', ''), record.get('username', '')))}\n"
            f"ID : <code>{record.get('user_id')}</code>\n"
            f"Raison : {reason}"
        )
        await update.message.reply_text(
            text,
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(
                profile_action_rows(record) + access_manage_keyboard(int(record.get("user_id")), pending=False, blacklisted=True).inline_keyboard
            ),
        )


async def history_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_admin(update):
        return
    history = list(ACCESS_STATE.get("history", []))[-15:]
    if not history:
        await update.message.reply_text("Aucun historique d'accès.", parse_mode=ParseMode.HTML)
        return
    lines = ["🕘 <b>Historique accès</b>", ""]
    for item in reversed(history):
        dt = datetime.fromtimestamp(int(item.get("ts") or now_ts())).strftime("%d/%m %H:%M")
        label = user_display_name(item.get("first_name", ""), item.get("username", ""))
        lines.append(f"• {dt} — <b>{esc(item.get('event', ''))}</b> — {esc(label)} — <code>{item.get('user_id')}</code>")
    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)

# --------------------------------------------------
# PROSPECTS
# --------------------------------------------------
def get_keyword_aliases(keyword: str) -> List[str]:
    normalized = normalize_keyword(keyword)
    lower = normalized.lower()

    alias_map = {
        "rh": [
            "rh", "ressources humaines", "human resources", "hr",
            "talent acquisition", "recrutement", "charge de recrutement",
            "responsable recrutement"
        ],
        "hr": [
            "hr", "human resources", "ressources humaines",
            "talent acquisition", "recrutement"
        ],
        "rrh": ["rrh", "responsable ressources humaines", "human resources manager"],
        "drh": ["drh", "directeur ressources humaines", "directrice ressources humaines", "hr director"],
        "commercial": [
            "commercial", "sales", "business development", "bizdev",
            "account executive", "charge d affaires", "ingénieur d affaires",
            "ingenieur d affaires"
        ],
        "marketing": [
            "marketing", "growth", "brand", "digital marketing",
            "content marketing", "communication", "marketing digital"
        ],
        "finance": ["finance", "financial", "comptabilite", "accounting", "controle de gestion"],
        "data": ["data", "data analyst", "data scientist", "business intelligence", "bi"],
        "recruteur": ["recruteur", "recruiter", "talent acquisition", "charge de recrutement"],
    }

    if lower in alias_map:
        return alias_map[lower]

    return [normalized]


def contains_alias(text: str, aliases: List[str]) -> bool:
    haystack = normalize_text(text)
    if not haystack:
        return False

    for alias in aliases:
        alias = normalize_text(alias)
        if not alias:
            continue

        if len(alias) <= 3 and " " not in alias:
            pattern = rf"\b{re.escape(alias)}\b"
            if re.search(pattern, haystack, flags=re.IGNORECASE):
                return True
        else:
            if alias in haystack:
                return True

    return False


def row_matches_keyword(row: Dict[str, str], keyword: str) -> bool:
    aliases = get_keyword_aliases(keyword)

    poste = row.get("Poste", "") or ""
    entreprise = row.get("Entreprise", "") or ""
    snippet = row.get("Snippet", "") or ""

    strong_text = " ".join([poste, entreprise, snippet]).strip()
    return contains_alias(strong_text, aliases)


def build_prospect_query(keyword: str, custom_filters: Dict[str, str]) -> str:
    keyword = normalize_keyword(keyword)
    normalized_keyword = normalize_text(keyword)

    aliases = get_keyword_aliases(keyword)
    
    keyword_tokens = [t for t in normalized_keyword.split() if len(t) >= 2]

    if len(aliases) > 1:
        keyword_part = "(" + " OR ".join(f'"{a}"' for a in aliases[:6]) + ")"
    elif len(keyword_tokens) >= 2:
        keyword_part = " ".join(f'"{t}"' for t in keyword_tokens[:5])
    elif aliases:
        keyword_part = f'"{aliases[0]}"'
    else:
        keyword_part = f'"{keyword}"'

    query = (
        f'site:linkedin.com/in {keyword_part} '
        f'-jobs -job -hiring -recruitment -recrutement '
        f'-offres -offre -emploi -stage -alternance -apprentissage '
        f'-learning -formation -formations -posts -school -ecole -universite'
    )

    entreprise = clean_spaces(custom_filters.get("entreprise", ""))
    ville = clean_spaces(custom_filters.get("ville", ""))
    pays = clean_spaces(custom_filters.get("pays", ""))
    secteur = clean_spaces(custom_filters.get("secteur", ""))
    poste = clean_spaces(custom_filters.get("poste", ""))
    seniorite = clean_spaces(custom_filters.get("seniorite", ""))
    region_scope = clean_spaces(custom_filters.get("region_scope", "")).lower()

    if poste:
        query += f' "{poste}"'
    if entreprise:
        query += f' "{entreprise}"'
    if ville:
        query += f' "{ville}"'
    if pays:
        query += f' "{pays}"'
    if secteur:
        query += f' "{secteur}"'
    if seniorite:
        query += f' "{seniorite}"'

    if region_scope == "france":
        query += " (\"France\" OR \"Paris\" OR \"Lyon\" OR \"Lille\" OR \"Marseille\" OR \"Toulouse\" OR \"Nantes\")"
    elif region_scope == "maghreb":
        query += " (\"Maroc\" OR \"Morocco\" OR \"Tunisie\" OR \"Tunisia\" OR \"Algérie\" OR \"Algeria\")"
    elif region_scope == "europe":
        query += " (\"France\" OR \"Germany\" OR \"Allemagne\" OR \"Spain\" OR \"Espagne\" OR \"Italy\" OR \"Italie\" OR \"Belgium\" OR \"Belgique\" OR \"Netherlands\" OR \"Pays-Bas\")"

    return query

def is_job_offer(text: str) -> bool:
    text = normalize_text(text)

    keywords = [
        "recherche", "poste", "mission", "profil", "experience",
        "responsable", "manager", "recrutons", "offre", "job",
        "role", "position", "cdi", "cdd", "alternance", "stage",
        "candidat", "entreprise", "client", "competences", "qualification",
        "fullstack", "backend", "frontend", "developpeur", "developer",
        "react", "node", "python", "java", "data", "sales", "commercial"
    ]

    return any(k in text for k in keywords)


def extract_job_params(job_text: str) -> Dict[str, Any]:
    raw_text = clean_spaces(job_text)
    text = normalize_text(raw_text)

    words = text.split()
    text = " ".join(words[:50])

    job_titles: List[str] = []
    keywords: List[str] = []
    experience_years = 0
    

    if any(x in text for x in ["sales", "commercial", "business developer", "bizdev", "account executive"]):
        job_titles += ["sales manager", "account executive", "business developer", "commercial"]

    if any(x in text for x in ["marketing", "growth", "brand", "acquisition"]):
        job_titles += ["marketing manager", "growth manager", "digital marketing"]

    if any(x in text for x in ["data analyst", "bi", "business intelligence"]):
        job_titles += ["data analyst", "business intelligence"]

    if any(x in text for x in ["data scientist", "machine learning", "ml"]):
        job_titles += ["data scientist"]

    if any(x in text for x in ["fullstack", "full stack"]):
        job_titles += ["developpeur fullstack"]

    if any(x in text for x in ["backend", "back end", "back-end"]):
        job_titles += ["developpeur backend"]

    if any(x in text for x in ["frontend", "front end", "front-end"]):
        job_titles += ["developpeur frontend"]

    if any(x in text for x in ["devops", "kubernetes", "docker", "ci cd", "ci/cd"]):
        job_titles += ["devops engineer"]

    tech_map = [
        "react", "node", "nodejs", "python", "java", "spring", "sql",
        "power bi", "tableau", "aws", "azure", "gcp", "kubernetes",
        "docker", "fastapi", "django", "salesforce", "hubspot",
        "excel", "sap", "crm", "saas", "b2b", "devops"
    ]
    for tech in tech_map:
        if normalize_text(tech) in text:
            keywords.append(tech)

    exp_match = re.search(r"(\d+)\s*(ans|years)", text)
    if exp_match:
        experience_years = int(exp_match.group(1))

    city_match = re.search(
        r"\b(paris|lyon|marseille|lille|bordeaux|toulouse|nantes|nice|cannes|rennes|strasbourg|montpellier)\b",
        text
    )
    city = city_match.group(1) if city_match else ""

    seniority = ""
    if any(x in text for x in ["lead", "principal", "head of"]):
        seniority = "lead"
    elif any(x in text for x in ["senior", "confirme", "confirmé", "expert"]):
        seniority = "senior"
    elif any(x in text for x in ["junior", "debutant", "débutant"]):
        seniority = "junior"

    dedup_titles = []
    seen_titles = set()
    for t in job_titles:
        nt = normalize_text(t)
        if nt and nt not in seen_titles:
            seen_titles.add(nt)
            dedup_titles.append(t)

    dedup_keywords = []
    seen_keywords = set()
    for k in keywords:
        nk = normalize_text(k)
        if nk and nk not in seen_keywords:
            seen_keywords.add(nk)
            dedup_keywords.append(k)

    if not dedup_titles:
        strong_words = [w for w in words if len(w) > 3][:4]
        dedup_titles = [" ".join(strong_words)] if strong_words else [raw_text]

    return {
        "job_titles": dedup_titles[:6],
        "keywords": dedup_keywords[:6],
        "city": city,
        "experience": experience_years,
        "seniority": seniority,
    }


def build_multi_queries(params: Dict[str, Any]) -> List[str]:
    queries = []
    
    for title in params.get("job_titles", []):
        parts = [title]
        parts.extend(params.get("keywords", []))
        if params.get("city"):
            parts.append(params["city"])
        base = clean_spaces(" ".join(parts))
        if base:
            queries.append(base)
            
    if not queries and params.get("job_titles"):
        queries = params["job_titles"][:]
        
    return list(dict.fromkeys(queries))[:MAX_PROSPECT_QUERY_VARIANTS]

def extract_experience_from_text(text: str) -> int:
    match = re.search(r"(\d+)\s*(ans|years)", text.lower())
    return int(match.group(1)) if match else 0

def row_matches_job_offer(row: Dict[str, str], params: Dict[str, Any]) -> bool:
    text = normalize_text(" ".join([
        row.get("Poste", ""),
        row.get("Entreprise", ""),
        row.get("Snippet", ""),
    ]))

    title_hit = False
    for t in params.get("job_titles", []):
        nt = normalize_text(t)
        if nt and nt in text:
            title_hit = True
            break

    keyword_hits = 0
    for k in params.get("keywords", []):
        nk = normalize_text(k)
        if nk and nk in text:
            keyword_hits += 1

    if title_hit:
        return True
    if keyword_hits >= 1:
        return True

    return False


def score_profile_advanced(row: Dict[str, str], params: Dict[str, Any], custom_filters: Dict[str, str]) -> int:
    score = 0

    title = normalize_text(row.get("Poste", ""))
    snippet = normalize_text(row.get("Snippet", ""))
    combined = f"{title} {snippet}"

    # 🎯 TITRE
    for t in params.get("job_titles", []):
        nt = normalize_text(t)
        if nt in title:
            score += 50
            break
        elif nt in combined:
            score += 25

    # 🎯 KEYWORDS
    for k in params.get("keywords", []):
        nk = normalize_text(k)
        if nk in combined:
            score += 15

    # 🎯 VILLE
    if params.get("city") and params["city"] in combined:
        score += 20

    # 🔥 EXPERIENCE
    required_exp = params.get("experience", 0)
    profile_exp = extract_experience_from_text(snippet)

    if required_exp:
        if profile_exp >= required_exp:
            score += 25
        else:
            score -= 10

    return score

# --------------------------------------------------
# RECHERCHE PERSONNE LINKEDIN
# --------------------------------------------------
def split_name(full_name: str) -> List[str]:
    return [p for p in normalize_text(full_name).split() if p]


def generate_first_name_variants(first_name: str) -> List[str]:
    first = normalize_text(first_name)
    variants = {first}

    custom_map = {
        "sarah": {"sarah", "sara", "sarra"},
        "sarra": {"sarra", "sara", "sarah"},
        "sara": {"sara", "sarah", "sarra"},
        "mohamed": {"mohamed", "mohammed", "muhammad", "mohamad"},
        "mohammed": {"mohammed", "mohamed", "muhammad", "mohamad"},
        "amina": {"amina", "aminah"},
        "aminah": {"aminah", "amina"},
    }

    if first in custom_map:
        variants.update(custom_map[first])

    if first.endswith("ah"):
        variants.add(first[:-1])
    if first.endswith("a"):
        variants.add(first + "h")
    if "rr" in first:
        variants.add(first.replace("rr", "r"))
    if "r" in first and "rr" not in first:
        variants.add(first.replace("r", "rr", 1))

    return [v for v in variants if v]


def build_person_query(person_name: str, custom_filters: Dict[str, str], fuzzy_enabled: bool) -> str:
    parts = split_name(person_name)

    if not parts:
        return ""

    first = parts[0]
    last = parts[-1]

    name_variants = []

    if fuzzy_enabled:
        variants = generate_first_name_variants(first)
        for v in variants:
            name_variants.append(f'"{v} {last}"')
    else:
        name_variants.append(f'"{first} {last}"')

    name_query = "(" + " OR ".join(name_variants) + ")"
    query = (
        f"site:linkedin.com/in {name_query} "
        f"-jobs -job -hiring -recruitment -recrutement "
        f"-offres -offre -emploi -stage -alternance -apprentissage "
        f"-posts -school -ecole -universite"
    )

    entreprise = custom_filters.get("entreprise")
    ville = custom_filters.get("ville")
    pays = custom_filters.get("pays")

    if entreprise:
        query += f' "{entreprise}"'
    if ville:
        query += f' "{ville}"'
    if pays:
        query += f' "{pays}"'

    return query


def compute_person_match_score(result_name: str, target_name: str, fuzzy_enabled: bool) -> float:
    result_parts = split_name(result_name)
    target_parts = split_name(target_name)

    if not result_parts or not target_parts:
        return 0.0

    result_full = " ".join(result_parts)
    target_full = " ".join(target_parts)
    full_ratio = SequenceMatcher(None, result_full, target_full).ratio()

    if not fuzzy_enabled:
        return full_ratio if result_full == target_full else 0.0

    result_first = result_parts[0]
    target_first = target_parts[0]
    result_last = result_parts[-1]
    target_last = target_parts[-1]

    first_ratio = SequenceMatcher(None, result_first, target_first).ratio()
    last_ratio = SequenceMatcher(None, result_last, target_last).ratio()

    return (first_ratio * 0.45) + (last_ratio * 0.45) + (full_ratio * 0.10)


def row_matches_person(row: Dict[str, str], target_name: str, fuzzy_enabled: bool) -> bool:
    score = compute_person_match_score(row.get("Nom", ""), target_name, fuzzy_enabled)
    row["MatchScore"] = round(score, 2)

    if fuzzy_enabled:
        return score >= 0.65
    return score >= 0.80

# --------------------------------------------------
# ENTREPRISES - OFFICIEL + ENRICHISSEMENT
# --------------------------------------------------
def build_annuaire_search_params(name: str, ville_filter: str = "", page: int = 1, per_page: int = 10) -> Dict[str, str]:
    q = clean_spaces(f"{name} {ville_filter}".strip())
    return {
        "q": q,
        "page": str(page),
        "per_page": str(per_page),
        "etat_administratif": "A",
    }


def build_annuaire_source_link(siren: str) -> str:
    siren = re.sub(r"\D", "", siren or "")
    if len(siren) == 9:
        return f"https://annuaire-entreprises.data.gouv.fr/entreprise/{siren}"
    return ""


def format_dirigeant_name(dirigeant: dict) -> str:
    if not isinstance(dirigeant, dict):
        return ""
    return clean_spaces(" ".join([
        str(dirigeant.get("prenoms", "") or ""),
        str(dirigeant.get("nom", "") or ""),
    ])).strip()


def extract_annuaire_dirigeants(item: dict) -> str:
    dirigeants = item.get("dirigeants") or []
    if not isinstance(dirigeants, list):
        return ""

    names = []
    for d in dirigeants[:3]:
        name = format_dirigeant_name(d)
        if name:
            qualite = clean_spaces(str(d.get("qualite", "") or ""))
            if qualite:
                names.append(f"{name} ({qualite})")
            else:
                names.append(name)

    return " ; ".join(names)


def extract_annuaire_activity(item: dict) -> str:
    return first_nonempty(
        item.get("libelle_activite_principale"),
        item.get("activite_principale"),
    )


def extract_annuaire_city(item: dict) -> str:
    siege = item.get("siege") or {}
    return first_nonempty(
        siege.get("libelle_commune"),
        item.get("libelle_commune"),
    )


def extract_annuaire_address(item: dict) -> str:
    siege = item.get("siege") or {}
    return first_nonempty(
        siege.get("adresse_complete"),
        clean_spaces(" ".join([
            str(siege.get("numero_voie", "") or ""),
            str(siege.get("type_voie", "") or ""),
            str(siege.get("libelle_voie", "") or ""),
            str(siege.get("code_postal", "") or ""),
            str(siege.get("libelle_commune", "") or ""),
        ])),
    )


def extract_annuaire_company_name(item: dict) -> str:
    return first_nonempty(
        item.get("nom_complet"),
        item.get("nom_raison_sociale"),
        item.get("denomination"),
        item.get("sigle"),
    )


def extract_annuaire_creation_date(item: dict) -> str:
    return first_nonempty(item.get("date_creation"))


def compute_company_info_score(row: Dict[str, str]) -> int:
    fields = [
        "Entreprise_INPI",
        "SIREN",
        "Dirigeant",
        "Ville_INPI",
        "Adresse_INPI",
        "Lien_source",
        "Snippet_entreprise",
        "Activite",
        "Date_creation",
    ]
    score = 0
    for field in fields:
        if (row.get(field, "") or "").strip():
            score += 1
    return score


def compute_company_relevance(row: Dict[str, str], target_name: str, ville_filter: str = "") -> Tuple[int, int]:
    parts = split_name(target_name)
    full_name = " ".join(parts).strip()
    last_name = parts[-1] if parts else target_name.strip()

    searchable_text = " ".join([
        row.get("Entreprise_INPI", "") or "",
        row.get("Dirigeant", "") or "",
        row.get("Ville_INPI", "") or "",
        row.get("Adresse_INPI", "") or "",
        row.get("Snippet_entreprise", "") or "",
        row.get("Lien_source", "") or "",
    ])

    haystack = normalize_text(searchable_text)
    ville_norm = normalize_text(ville_filter or "")
    full_name_norm = normalize_text(full_name)
    last_name_norm = normalize_text(last_name)

    full_name_match = bool(full_name_norm and full_name_norm in haystack)
    last_name_match = bool(last_name_norm and re.search(rf"\b{re.escape(last_name_norm)}\b", haystack))
    ville_match = bool(ville_norm and ville_norm in haystack)

    if full_name_match and ville_match:
        relevance = 3
        label = "Nom + prénom + ville"
    elif full_name_match:
        relevance = 2
        label = "Nom + prénom"
    elif last_name_match:
        relevance = 1
        label = "Nom"
    else:
        relevance = 0
        label = "Approx"

    info_score = compute_company_info_score(row)
    row["RelevanceLabel"] = label
    row["InfoScore"] = info_score

    return relevance, info_score


def annuaire_item_to_row(item: dict, query_name: str, ville_filter: str = "") -> Dict[str, str]:
    company_name = extract_annuaire_company_name(item)
    siren = first_nonempty(item.get("siren"))
    city = extract_annuaire_city(item)
    address = extract_annuaire_address(item)
    dirigeant = extract_annuaire_dirigeants(item)
    activite = extract_annuaire_activity(item)
    date_creation = extract_annuaire_creation_date(item)

    row = {
        "Entreprise_INPI": company_name or "N/A",
        "SIREN": siren or "",
        "Ville_INPI": city or "",
        "Adresse_INPI": address or "",
        "Dirigeant": dirigeant or "",
        "Activite": activite or "",
        "Date_creation": date_creation or "",
        "Source_entreprise": "annuaire_api",
        "Lien_source": build_annuaire_source_link(siren),
        "Snippet_entreprise": first_nonempty(
            activite,
            address,
            city,
            company_name,
        ),
    }

    relevance, info_score = compute_company_relevance(row, query_name, ville_filter)
    row["RelevanceScore"] = relevance
    row["InfoScore"] = info_score
    return row


def search_company_annuaire(name: str, ville_filter: str = "", max_results: int = 12) -> List[Dict[str, str]]:
    cache_key = f"annuaire__{normalize_text(name)}__{normalize_text(ville_filter)}__{max_results}"
    cached = get_cache_payload(cache_key)
    if cached:
        return cached.get("results", [])

    headers = {
        "Accept": "application/json",
        "User-Agent": ANNUAIRE_USER_AGENT,
    }

    params = build_annuaire_search_params(name, ville_filter, page=1, per_page=max_results)

    response = requests.get(
        f"{ANNUAIRE_API_BASE}/search",
        headers=headers,
        params=params,
        timeout=REQUEST_TIMEOUT_TUPLE,
    )
    response.raise_for_status()
    data = response.json()

    rows = []
    for item in data.get("results", []) or []:
        if isinstance(item, dict):
            rows.append(annuaire_item_to_row(item, name, ville_filter))

    rows.sort(
        key=lambda r: (
            safe_int(r.get("RelevanceScore", 0)),
            safe_int(r.get("InfoScore", 0)),
            len((r.get("Dirigeant", "") or "").strip()),
            len((r.get("Adresse_INPI", "") or "").strip()),
        ),
        reverse=True,
    )

    PROVIDER_STATS["annuaire_api"] += 1
    payload = {"results": rows[:max_results]}
    set_cache_payload(cache_key, payload)
    return payload["results"]


def build_company_queries(name: str, ville_filter: str = "") -> List[str]:
    parts = split_name(name)
    exact_name = " ".join(parts).strip()
    last_name = parts[-1] if parts else name.strip()

    base_queries: List[str] = []

    if exact_name:
        if ville_filter:
            base_queries.extend([
                f'site:annuaire-entreprises.data.gouv.fr/entreprise "{exact_name}" "{ville_filter}"',
                f'site:societe.com "{exact_name}" "{ville_filter}"',
                f'site:pappers.fr "{exact_name}" "{ville_filter}"',
                f'site:verif.com "{exact_name}" "{ville_filter}"',
                f'site:annuaire-entreprises.data.gouv.fr "{exact_name}" dirigeant "{ville_filter}"',
            ])
        else:
            base_queries.extend([
                f'site:annuaire-entreprises.data.gouv.fr/entreprise "{exact_name}"',
                f'site:societe.com "{exact_name}"',
                f'site:pappers.fr "{exact_name}"',
                f'site:verif.com "{exact_name}"',
                f'site:annuaire-entreprises.data.gouv.fr "{exact_name}" dirigeant',
                f'site:societe.com "{exact_name}" dirigeant',
            ])

    if last_name and last_name.lower() != exact_name.lower():
        if ville_filter:
            base_queries.extend([
                f'site:annuaire-entreprises.data.gouv.fr/entreprise "{last_name}" "{ville_filter}"',
                f'site:societe.com "{last_name}" "{ville_filter}"',
            ])
        else:
            base_queries.extend([
                f'site:annuaire-entreprises.data.gouv.fr/entreprise "{last_name}"',
                f'site:societe.com "{last_name}"',
            ])

    return list(dict.fromkeys(base_queries))[:MAX_COMPANY_QUERY_VARIANTS]


def is_company_domain(link: str) -> bool:
    if not link:
        return False
    link = link.lower()
    return any(domain in link for domain in [
        "annuaire-entreprises.data.gouv.fr",
        "societe.com",
        "pappers.fr",
        "manageo.fr",
        "verif.com",
    ])


def extract_siren(text: str) -> str:
    if not text:
        return ""
    cleaned = text.replace(" ", "").replace(".", "")
    match = re.search(r"\b(\d{9})\b", cleaned)
    return match.group(1) if match else ""


def extract_city_from_text(text: str) -> str:
    if not text:
        return ""
    match = re.search(r"\b\d{5}\s+([A-Za-zÀ-ÿ\-\s']{2,40})", text)
    if match:
        return match.group(1).strip(" ,.-")
    return ""


def extract_address_from_text(text: str) -> str:
    if not text:
        return ""
    match = re.search(r"([0-9]{1,4}[^.:\n]{5,160}\b\d{5}\s+[A-Za-zÀ-ÿ\-\s']{2,40})", text)
    if match:
        return match.group(1).strip(" ,.-")
    return ""


def extract_company_name_from_title(title: str) -> str:
    if not title:
        return ""

    title = title.strip()
    patterns_to_remove = [
        " - Annuaire des entreprises",
        "| Annuaire des entreprises",
        "Annuaire des entreprises",
        " - Société.com",
        "| Société.com",
        " - Pappers",
        "| Pappers",
        " - Manageo.fr",
        "| Manageo.fr",
        " - Manageo",
        "| Manageo",
        " - Verif.com",
        "| Verif.com",
        " - Verif",
        "| Verif",
    ]

    for p in patterns_to_remove:
        title = title.replace(p, "")

    return re.sub(r"\s+", " ", title).strip(" -|")


def extract_dirigeant_from_text(text: str) -> str:
    if not text:
        return ""

    text = re.sub(r"\s+", " ", text)

    patterns = [
        r"(?:dirigeant|gérant|gerant|président|president|représentant légal|representant legal|présidente|presidente|président du conseil|chef d entreprise)\s*[:\-]?\s*([A-ZÀ-Ÿ][A-Za-zÀ-ÿ'\-]+\s+[A-ZÀ-Ÿ][A-Za-zÀ-ÿ'\-]+)",
        r"([A-ZÀ-Ÿ][A-Za-zÀ-ÿ'\-]+\s+[A-ZÀ-Ÿ][A-Za-zÀ-ÿ'\-]+)\s*(?:dirigeant|gérant|gerant|président|president|présidente|presidente)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()

    return ""


def count_strong_company_results(rows: List[Dict[str, str]]) -> int:
    count = 0
    for row in rows:
        if safe_int(row.get("RelevanceScore", 0)) >= 2 and safe_int(row.get("InfoScore", 0)) >= 5:
            count += 1
    return count


def search_company_web_query(query: str, name: str, ville_filter: str = "") -> List[Dict[str, str]]:
    results: List[Dict[str, str]] = []
    ville_filter_norm = normalize_text(ville_filter or "")

    data = search_web(query, start=0, num=10)
    organic_results = data.get("organic_results", [])

    for item in organic_results:
        link = (item.get("link", "") or "").strip()
        title = (item.get("title", "") or "").strip()
        snippet = (item.get("snippet", "") or "").strip()

        if not link:
            continue
        if not is_company_domain(link):
            continue

        combined_text = " ".join([title, snippet, link])

        row = {
            "Entreprise_INPI": extract_company_name_from_title(title) or title[:120] or "N/A",
            "SIREN": extract_siren(combined_text),
            "Ville_INPI": extract_city_from_text(combined_text),
            "Adresse_INPI": extract_address_from_text(combined_text),
            "Dirigeant": extract_dirigeant_from_text(combined_text),
            "Activite": "",
            "Date_creation": "",
            "Source_entreprise": data.get("provider", "web"),
            "Lien_source": link,
            "Snippet_entreprise": snippet,
        }

        if ville_filter_norm:
            haystack = normalize_text(
                f"{row.get('Ville_INPI', '')} {row.get('Adresse_INPI', '')} {title} {snippet}"
            )
            if ville_filter_norm not in haystack:
                continue

        relevance, info_score = compute_company_relevance(row, name, ville_filter)
        row["RelevanceScore"] = relevance
        row["InfoScore"] = info_score
        results.append(row)

    return results


def search_company_web_enrichment(name: str, ville_filter: str = "") -> List[Dict[str, str]]:
    queries = build_company_queries(name, ville_filter)[:COMPANY_WEB_QUERY_BUDGET]
    if len(queries) >= COMPANY_WEB_QUERY_BUDGET:
        PROVIDER_STATS["company_query_budget_hits"] += 1
    collected: List[Dict[str, str]] = []
    seen_keys = set()

    max_workers = min(COMPANY_WEB_MAX_WORKERS, len(queries)) or 1

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(search_company_web_query, q, name, ville_filter): q
            for q in queries
        }

        for future in as_completed(futures):
            try:
                rows = future.result()
            except Exception as e:
                logger.warning("Erreur enrichissement requête web: %s", e)
                continue

            for row in rows:
                siren = (row.get("SIREN", "") or "").strip()
                name_key = normalize_text(row.get("Entreprise_INPI", "") or "")
                city_key = normalize_text(row.get("Ville_INPI", "") or "")
                key = siren or f"{name_key}__{city_key}"

                if key in seen_keys:
                    continue

                seen_keys.add(key)
                collected.append(row)

                if len(collected) >= COMPANY_WEB_MAX_ENRICH_CANDIDATES:
                    break

            if count_strong_company_results(collected) >= MAX_COMPANY_STRONG_RESULTS:
                break
            if len(collected) >= COMPANY_WEB_MAX_ENRICH_CANDIDATES:
                break

    collected.sort(
        key=lambda r: (
            safe_int(r.get("RelevanceScore", 0)),
            safe_int(r.get("InfoScore", 0)),
            len((r.get("Dirigeant", "") or "").strip()),
            len((r.get("Adresse_INPI", "") or "").strip()),
        ),
        reverse=True,
    )

    return collected[:COMPANY_WEB_MAX_ENRICH_CANDIDATES]


def merge_company_sources(primary_rows: List[Dict[str, str]], secondary_rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    merged: List[Dict[str, str]] = []
    seen_keys = set()

    for row in primary_rows + secondary_rows:
        siren = (row.get("SIREN", "") or "").strip()
        name_key = normalize_text(row.get("Entreprise_INPI", "") or "")
        city_key = normalize_text(row.get("Ville_INPI", "") or "")
        key = siren or f"{name_key}__{city_key}"

        if key in seen_keys:
            continue

        seen_keys.add(key)
        merged.append(row)

    merged.sort(
        key=lambda r: (
            safe_int(r.get("RelevanceScore", 0)),
            safe_int(r.get("InfoScore", 0)),
            1 if r.get("Source_entreprise") == "annuaire_api" else 0,
            len((r.get("Dirigeant", "") or "").strip()),
            len((r.get("Adresse_INPI", "") or "").strip()),
        ),
        reverse=True,
    )

    return merged[:MAX_COMPANY_TOTAL_RESULTS]


def search_company_person(name: str, ville_filter: str = "") -> List[Dict[str, str]]:
    cache_key = f"company__{normalize_cache_phrase(name)}__{normalize_cache_phrase(ville_filter)}"
    cached = get_cache_payload(cache_key)
    if cached:
        return cached.get("results", [])

    annuaire_rows: List[Dict[str, str]] = []
    web_rows: List[Dict[str, str]] = []

    try:
        annuaire_rows = search_company_annuaire(name, ville_filter, max_results=12)
    except Exception as e:
        logger.warning("Annuaire API indisponible: %s", e)

    annuaire_rows = sorted(
        annuaire_rows,
        key=lambda r: (
            safe_int(r.get("RelevanceScore", 0)),
            safe_int(r.get("InfoScore", 0)),
        ),
        reverse=True,
    )

    if count_strong_company_results(annuaire_rows) < MAX_COMPANY_STRONG_RESULTS:
        try:
            web_rows = search_company_web_enrichment(name, ville_filter)
        except Exception as e:
            logger.warning("Enrichissement web indisponible: %s", e)

    final_rows = merge_company_sources(annuaire_rows, web_rows)
    payload = {"results": final_rows[:MAX_COMPANY_TOTAL_RESULTS], "negative_cache": len(final_rows) == 0}
    set_cache_payload(cache_key, payload)
    return payload["results"]

# --------------------------------------------------
# AFFICHAGE ENTREPRISE
# --------------------------------------------------
def build_company_summary_line(global_index: int, row: Dict[str, str]) -> str:
    entreprise = esc(row.get("Entreprise_INPI", "") or "N/A")
    siren = esc(row.get("SIREN", "") or "N/A")
    pertinence = esc(row.get("RelevanceLabel", "") or "N/A")
    dirigeant = esc(row.get("Dirigeant", "") or "N/A")
    ville = esc(row.get("Ville_INPI", "") or "N/A")
    source = esc(row.get("Source_entreprise", "") or "N/A")

    return (
        f"<b>{global_index}. {entreprise}</b>\n"
        f"• <b>SIREN :</b> {siren}\n"
        f"• <b>Pertinence :</b> {pertinence}\n"
        f"• <b>Dirigeant :</b> {dirigeant}\n"
        f"• <b>Ville :</b> {ville}\n"
        f"• <b>Source :</b> {source}"
    )


def build_company_detail_text(index: int, row: Dict[str, str]) -> str:
    return (
        f"<b>Détails du résultat {index}</b>\n\n"
        f"• <b>Entreprise :</b> {esc(row.get('Entreprise_INPI', '') or 'N/A')}\n"
        f"• <b>SIREN :</b> {esc(row.get('SIREN', '') or 'N/A')}\n"
        f"• <b>Pertinence :</b> {esc(row.get('RelevanceLabel', '') or 'N/A')}\n"
        f"• <b>Dirigeant :</b> {esc(row.get('Dirigeant', '') or 'N/A')}\n"
        f"• <b>Adresse :</b> {esc(row.get('Adresse_INPI', '') or 'N/A')}\n"
        f"• <b>Ville :</b> {esc(row.get('Ville_INPI', '') or 'N/A')}\n"
        f"• <b>Activité :</b> {esc(row.get('Activite', '') or 'N/A')}\n"
        f"• <b>Date création :</b> {esc(row.get('Date_creation', '') or 'N/A')}\n"
        f"• <b>Source :</b> {esc(row.get('Source_entreprise', '') or 'N/A')}\n"
        f"• <b>Lien :</b> {esc(row.get('Lien_source', '') or 'N/A')}\n\n"
        f"<b>Snippet</b>\n{esc(row.get('Snippet_entreprise', '') or 'N/A')}"
    )


def build_company_page_text(companies: List[Dict[str, str]], current_page: int) -> str:
    start_index = current_page * COMPANY_PAGE_SIZE
    subset = companies[start_index:start_index + COMPANY_PAGE_SIZE]

    if not subset:
        return "Plus de résultats disponibles."

    lines = []
    for i, row in enumerate(subset, start=start_index + 1):
        lines.append(build_company_summary_line(i, row))

    total_pages = (len(companies) + COMPANY_PAGE_SIZE - 1) // COMPANY_PAGE_SIZE
    header = f"<b>Résultats entreprise</b> — page <b>{current_page + 1}/{total_pages}</b>\n\n"
    return header + "\n\n".join(lines)

# --------------------------------------------------
# PARSING TITRE GOOGLE
# --------------------------------------------------
def extract_years_experience(text: str) -> int:
    text = normalize_text(text)
    patterns = [
        r"(\d+)\s*ans",
        r"(\d+)\s*years",
        r"minimum\s*(\d+)",
        r"at least\s*(\d+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            try:
                return int(match.group(1))
            except Exception:
                pass
    return 0


def extract_seniority(text: str) -> str:
    text = normalize_text(text)

    mapping = {
        "lead": ["lead", "principal", "head of"],
        "senior": ["senior", "confirme", "confirmé", "expert"],
        "mid": ["intermediaire", "intermédiaire", "3 ans", "4 ans", "5 ans"],
        "junior": ["junior", "debutant", "débutant", "1 an", "2 ans"],
    }

    for level, keywords in mapping.items():
        for kw in keywords:
            if normalize_text(kw) in text:
                return level

    return ""


def extract_technologies(text: str) -> List[str]:
    text = normalize_text(text)

    techs = [
        "react", "node", "nodejs", "python", "java", "spring", "sql",
        "power bi", "tableau", "aws", "azure", "gcp", "kubernetes",
        "docker", "fastapi", "django", "salesforce", "hubspot",
        "excel", "sap", "crm", "saas", "b2b", "devops"
    ]

    found = []
    for tech in techs:
        if normalize_text(tech) in text:
            found.append(tech)

    return list(dict.fromkeys(found))

def smart_extract_search_query(job_text: str) -> Dict[str, Any]:
    text = normalize_text(job_text)
    words = text.split()
    text = " ".join(words[:80])

    job = ""
    city = ""
    keywords = extract_technologies(text)
    experience = extract_years_experience(text)
    seniority = extract_seniority(text)

    if any(x in text for x in ["fullstack", "full stack"]):
        job = "developpeur fullstack"
    elif any(x in text for x in ["backend", "back-end"]):
        job = "developpeur backend"
    elif any(x in text for x in ["frontend", "front-end"]):
        job = "developpeur frontend"
    elif any(x in text for x in ["data analyst", "bi", "business intelligence"]):
        job = "data analyst"
    elif any(x in text for x in ["data scientist"]):
        job = "data scientist"
    elif any(x in text for x in ["devops", "sre"]):
        job = "devops engineer"
    elif any(x in text for x in ["sales", "commercial", "business developer", "account executive"]):
        job = "commercial"
    elif any(x in text for x in ["marketing", "growth"]):
        job = "marketing manager"
    elif any(x in text for x in ["rh", "recrutement", "recruiter", "talent acquisition"]):
        job = "recruiter"
    else:
        strong_words = [w for w in words if len(w) > 3][:4]
        job = " ".join(strong_words) if strong_words else "profil"

    cities = [
        "paris", "lyon", "lille", "marseille", "bordeaux", "toulouse",
        "nantes", "nice", "cannes", "rennes", "strasbourg", "montpellier"
    ]
    for c in cities:
        if c in text:
            city = c
            break

    return {
        "job": job,
        "keywords": keywords[:5],
        "city": city,
        "experience": experience,
        "seniority": seniority,
    }
def parse_google_title(title: str) -> Dict[str, str]:
    cleaned = title.replace("| LinkedIn", "").strip()
    parts = [p.strip() for p in cleaned.split(" - ")]

    nom = parts[0] if len(parts) > 0 else ""
    poste = parts[1] if len(parts) > 1 else ""
    entreprise = parts[2] if len(parts) > 2 else ""

    return {
        "Nom": nom,
        "Poste": poste,
        "Entreprise": entreprise,
    }

# --------------------------------------------------
# EXCEL
# --------------------------------------------------
def export_excel(results: List[Dict[str, str]], keyword: str) -> str:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    Path(EXPORT_DIR).mkdir(exist_ok=True)
    safe_keyword = sanitize_filename(keyword)
    file_name = f"{EXPORT_DIR}/prospects_{safe_keyword}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    final_columns = [
        "Nom",
        "Poste",
        "Entreprise",
        "Entreprise_INPI",
        "SIREN",
        "Dirigeant",
        "Ville_INPI",
        "Adresse_INPI",
        "Ville",
        "Pays",
        "Source",
        "Statut",
        "Priorité",
        "MatchScore",
        "Notes",
        "Snippet",
        "LinkedIn",
        "Lien_source",
        "Snippet_entreprise",
    ]

    if not results:
        df = pd.DataFrame(columns=final_columns)
    else:
        df = pd.DataFrame(results)

    for col in final_columns:
        if col not in df.columns:
            if col == "Source":
                df[col] = "Google / LinkedIn"
            elif col == "Statut":
                df[col] = "À contacter"
            elif col == "Priorité":
                df[col] = "Moyenne"
            else:
                df[col] = ""

    df = df[final_columns].fillna("")

    for col in ["Nom", "Poste", "Entreprise", "Ville", "Pays", "LinkedIn"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    if "LinkedIn" in df.columns:
        df = df.drop_duplicates(subset=["LinkedIn"])
    if "Nom" in df.columns and "Entreprise" in df.columns:
        df = df.drop_duplicates(subset=["Nom", "Entreprise"], keep="first")

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Prospects")
        wb = writer.book
        ws = writer.sheets["Prospects"]

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 24

        ws.freeze_panes = "A2"

        last_row = ws.max_row
        last_col = ws.max_column
        if last_row >= 1 and last_col >= 1:
            table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
            table = Table(displayName="TableProspects", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        summary = wb.create_sheet("Résumé")
        summary["A1"] = "Recherche"
        summary["B1"] = keyword
        summary["A2"] = "Date export"
        summary["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        summary["A3"] = "Nombre de lignes"
        summary["B3"] = len(df)

        for cell_ref in ["A1", "A2", "A3"]:
            summary[cell_ref].font = Font(bold=True)

        summary.column_dimensions["A"].width = 22
        summary.column_dimensions["B"].width = 42

    return file_name

# --------------------------------------------------
# RECHERCHE PAGINÉE GÉNÉRIQUE
# --------------------------------------------------
def google_search_page(
    query: str,
    custom_filters: Dict[str, str],
    start: int,
    page_size: int,
) -> Dict[str, object]:
    del custom_filters
    result = search_web(query=query, start=start, num=page_size)

    return {
        "organic_results": result.get("organic_results", []),
        "has_more": bool(result.get("has_more", False)),
        "next_start": result.get("next_start", start + page_size),
    }

# --------------------------------------------------
# RECHERCHE PROSPECTS
# --------------------------------------------------
def search_prospect_page(
    keyword: str,
    custom_filters: Optional[Dict[str, str]] = None,
    start: int = 0,
    page_size: int = SERP_BATCH_SIZE,
) -> Dict[str, object]:
    custom_filters = custom_filters or {}
    cache_key = make_cache_key("prospect", keyword, custom_filters, start, page_size, False)
    cached = get_cache_payload(cache_key)
    if cached:
        return cached

    raw_keyword = clean_spaces(keyword)
    normalized_keyword = normalize_text(raw_keyword)
    smart = smart_extract_search_query(raw_keyword)

    # 🔥 beaucoup plus de variantes
    queries = list(dict.fromkeys([
        raw_keyword,
        smart["job"],
        f"{smart['job']} {' '.join(smart['keywords'])}" if smart["keywords"] else "",
        f"{smart['job']} freelance",
        f"{smart['job']} consultant",
        f"{smart['job']} developer",
        f"{smart['job']} engineer",
        f"{smart['job']} {' '.join(normalized_keyword.split()[:3])}",
    ]))[:10]

    queries = [q for q in queries if clean_spaces(q)]

    all_results = []
    seen_links = set()

    # 🔥 multi pages + multi queries
    for q in queries:
        for page_round in range(0, 2):  # 2 pages = ~60 résultats par query
            query = build_prospect_query(q, custom_filters)
            page = google_search_page(query, custom_filters, page_round * SERP_BATCH_SIZE, SERP_BATCH_SIZE)
            print("QUERY:", query)
            print("RESULTATS GOOGLE:", len(page.get("organic_results", [])))
            print("RESULTATS GOOGLE:", len(page.get("organic_results", [])))

            for item in page.get("organic_results", []):
                link = normalize_linkedin_url(item.get("link", ""))

                if not link or "linkedin.com/in/" not in link:
                    continue
                if link in seen_links:
                    continue

                seen_links.add(link)

                title_data = parse_google_title(item.get("title", ""))
                snippet = (item.get("snippet", "") or "").strip()

                row = {
                    "Nom": title_data.get("Nom", ""),
                    "Poste": title_data.get("Poste", ""),
                    "Entreprise": title_data.get("Entreprise", ""),
                    "Ville": custom_filters.get("ville", ""),
                    "Pays": custom_filters.get("pays", ""),
                    "Source": "Google / LinkedIn",
                    "Statut": "À contacter",
                    "Priorité": "Moyenne",
                    "Notes": "",
                    "Snippet": snippet,
                    "LinkedIn": link,
                }

                # 🔥 scoring seulement (PLUS DE FILTRE BLOQUANT)
                row["MatchScore"] = score_profile_advanced(row, {
                    "job_titles": [smart["job"]],
                    "keywords": smart["keywords"],
                    "city": smart["city"],
                    "experience": smart.get("experience", 0),
                    "seniority": smart.get("seniority", ""),
                }, custom_filters)

                all_results.append(row)

    # 🔥 tri uniquement (plus de suppression)
    all_results.sort(key=lambda x: x.get("MatchScore", 0), reverse=True)

    payload = {
        "results": all_results[start:start + page_size],
        "next_start": start + page_size,
        "has_more": len(all_results) > (start + page_size),
        "negative_cache": len(all_results) == 0,
    }
    set_cache_payload(cache_key, payload)
    return payload
# --------------------------------------------------
# RECHERCHE PERSONNE LINKEDIN
# -----------------------------------------------
def search_person_page(
    person_name: str,
    custom_filters: Optional[Dict[str, str]] = None,
    start: int = 0,
    page_size: int = SERP_BATCH_SIZE,
    fuzzy_enabled: bool = False,
) -> Dict[str, object]:
    custom_filters = custom_filters or {}
    cache_key = make_cache_key("person", person_name, custom_filters, start, page_size, fuzzy_enabled)

    cached = get_cache_payload(cache_key)
    if cached and not cached.get("negative_cache"):
        return cached

    query = build_person_query(person_name, custom_filters, fuzzy_enabled)
    page = google_search_page(query, custom_filters, start, page_size)
    organic_results = page["organic_results"]

    results: List[Dict[str, str]] = []
    seen_links = set()

    for item in organic_results:
        link = normalize_linkedin_url(item.get("link", ""))
        if not link or "linkedin.com/in/" not in link:
            continue
        if link in seen_links:
            continue
        seen_links.add(link)

        title_data = parse_google_title(item.get("title", ""))
        snippet = (item.get("snippet", "") or "").strip()

        row = {
            "Nom": title_data.get("Nom", ""),
            "Poste": title_data.get("Poste", ""),
            "Entreprise": title_data.get("Entreprise", ""),
            "Ville": custom_filters.get("ville", ""),
            "Pays": custom_filters.get("pays", ""),
            "Source": "Google / LinkedIn",
            "Statut": "À contacter",
            "Priorité": "Moyenne",
            "Notes": "",
            "Snippet": snippet,
            "LinkedIn": link,
        }

        if not row_matches_person(row, person_name, fuzzy_enabled):
            continue

        results.append(row)

        if start + len(results) >= MAX_RESULTS:
            break

    results.sort(key=lambda x: float(x.get("MatchScore", 0) or 0), reverse=True)

    payload = {
        "results": results,
        "next_start": page["next_start"],
        "has_more": bool(page["has_more"]) and (start + len(results) < MAX_RESULTS),
    }

    set_cache_payload(cache_key, payload)
    return payload

# --------------------------------------------------
# EXPORT COMPLET
# --------------------------------------------------
def search_full_export(
    search_mode: str,
    base_value: str,
    custom_filters: Optional[Dict[str, str]] = None,
    max_results: int = MAX_RESULTS,
    fuzzy_enabled: bool = False,
) -> List[Dict[str, str]]:
    custom_filters = custom_filters or {}
    all_results: List[Dict[str, str]] = []
    seen = set()
    start = 0

    while len(all_results) < max_results:
        if search_mode == "person":
            page = search_person_page(
                person_name=base_value,
                custom_filters=custom_filters,
                start=start,
                page_size=SERP_BATCH_SIZE,
                fuzzy_enabled=fuzzy_enabled,
            )
        else:
            page = search_prospect_page(
                keyword=base_value,
                custom_filters=custom_filters,
                start=start,
                page_size=SERP_BATCH_SIZE,
            )

        page_results = page.get("results", [])
        if not page_results:
            break

        for row in page_results:
            link = row.get("LinkedIn", "")
            if link and link not in seen:
                seen.add(link)
                all_results.append(row)
                if len(all_results) >= max_results:
                    break

        if not page.get("has_more", False):
            break

        start = page.get("next_start", start + SERP_BATCH_SIZE)

    return all_results

# --------------------------------------------------
# AFFICHAGE / PRÉCHARGEMENT
# --------------------------------------------------
async def fetch_display_chunk(
    search_mode: str,
    base_value: str,
    custom_filters: Dict[str, str],
    start: int,
    target_count: int = DISPLAY_PAGE_SIZE,
    fuzzy_enabled: bool = False,
) -> Dict[str, object]:
    collected: List[Dict[str, str]] = []
    current_start = start
    has_more_remote = True
    seen_links = set()

    while len(collected) < target_count and has_more_remote and len(collected) < MAX_RESULTS:
        if search_mode == "person":
            page = await asyncio.to_thread(
                search_person_page,
                base_value,
                custom_filters,
                current_start,
                SERP_BATCH_SIZE,
                fuzzy_enabled,
            )
        else:
            page = await asyncio.to_thread(
                search_prospect_page,
                base_value,
                custom_filters,
                current_start,
                SERP_BATCH_SIZE,
            )

        page_results = page.get("results", [])
        for row in page_results:
            link = row.get("LinkedIn", "")
            if link and link not in seen_links:
                seen_links.add(link)
                collected.append(row)
                if len(collected) >= target_count:
                    break

        has_more_remote = bool(page.get("has_more", False))
        current_start = page.get("next_start", current_start + SERP_BATCH_SIZE)

        if not page_results:
            break

    return {
        "results": collected[:target_count],
        "next_start": current_start,
        "has_more": has_more_remote,
    }


def build_result_blocks(results: List[Dict[str, str]], start_index: int, page_size: int = DISPLAY_PAGE_SIZE) -> List[str]:
    subset = results[start_index:start_index + page_size]
    blocks = []

    for idx, r in enumerate(subset, start=start_index + 1):
        nom = esc(r.get("Nom") or "Nom inconnu")
        poste = esc(r.get("Poste") or "N/A")
        entreprise = esc(r.get("Entreprise") or "N/A")
        lien = esc(r.get("LinkedIn") or "N/A")
        match_score = r.get("MatchScore", "")

        match_line = f"\n• <b>Match :</b> {esc(match_score)}" if str(match_score).strip() else ""

        block = (
            f"<b>{idx}. {nom}</b>\n"
            f"• <b>Poste :</b> {poste}\n"
            f"• <b>Entreprise :</b> {entreprise}"
            f"{match_line}\n"
            f"• <b>LinkedIn :</b> {lien}"
        )
        blocks.append(block)

    return blocks


def ensure_prefetch_state(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.setdefault("results", [])
    context.user_data.setdefault("index", 0)
    context.user_data.setdefault("next_start", 0)
    context.user_data.setdefault("has_more_remote", True)
    context.user_data.setdefault("prefetched_chunk", None)
    context.user_data.setdefault("prefetch_task", None)
    context.user_data.setdefault("page_lock", asyncio.Lock())


async def prefetch_next_chunk(context: ContextTypes.DEFAULT_TYPE) -> None:
    ensure_prefetch_state(context)

    if not context.user_data.get("has_more_remote", True):
        return

    if context.user_data.get("prefetched_chunk") is not None:
        return

    search_mode = context.user_data.get("search_mode")
    base_value = context.user_data.get("base_value")
    custom_filters = context.user_data.get("filters", {})
    start = context.user_data.get("next_start", 0)
    fuzzy_enabled = bool(context.user_data.get("fuzzy_enabled", False))

    if not search_mode or not base_value:
        return

    chunk = await fetch_display_chunk(
        search_mode=search_mode,
        base_value=base_value,
        custom_filters=custom_filters,
        start=start,
        target_count=DISPLAY_PAGE_SIZE,
        fuzzy_enabled=fuzzy_enabled,
    )

    context.user_data["prefetched_chunk"] = chunk
    context.user_data["has_more_remote"] = bool(chunk.get("has_more", False))


def start_prefetch(context: ContextTypes.DEFAULT_TYPE) -> None:
    ensure_prefetch_state(context)

    existing_task = context.user_data.get("prefetch_task")
    if existing_task and not existing_task.done():
        return

    if context.user_data.get("prefetched_chunk") is not None:
        return

    if not context.user_data.get("has_more_remote", True):
        return

    task = context.application.create_task(prefetch_next_chunk(context))
    context.user_data["prefetch_task"] = task


async def wait_for_prefetch_if_needed(context: ContextTypes.DEFAULT_TYPE) -> Optional[Dict[str, object]]:
    ensure_prefetch_state(context)

    cached_chunk = context.user_data.get("prefetched_chunk")
    if cached_chunk is not None:
        context.user_data["prefetched_chunk"] = None
        return cached_chunk

    task = context.user_data.get("prefetch_task")
    if task and not task.done():
        await task

    cached_chunk = context.user_data.get("prefetched_chunk")
    if cached_chunk is not None:
        context.user_data["prefetched_chunk"] = None
        return cached_chunk

    if context.user_data.get("has_more_remote", True):
        chunk = await fetch_display_chunk(
            search_mode=context.user_data.get("search_mode"),
            base_value=context.user_data.get("base_value"),
            custom_filters=context.user_data.get("filters", {}),
            start=context.user_data.get("next_start", 0),
            target_count=DISPLAY_PAGE_SIZE,
            fuzzy_enabled=bool(context.user_data.get("fuzzy_enabled", False)),
        )
        context.user_data["has_more_remote"] = bool(chunk.get("has_more", False))
        return chunk

    return None


async def send_next_page(chat_id: int, context: ContextTypes.DEFAULT_TYPE) -> None:
    ensure_prefetch_state(context)

    results = context.user_data.get("results", [])
    index = context.user_data.get("index", 0)

    if index >= len(results):
        chunk = await wait_for_prefetch_if_needed(context)

        if not chunk or not chunk.get("results"):
            await context.bot.send_message(
                chat_id=chat_id,
                text="Plus de résultats disponibles.",
                parse_mode=ParseMode.HTML,
            )
            return

        chunk_results = chunk["results"]
        context.user_data["results"].extend(chunk_results)
        context.user_data["next_start"] = chunk["next_start"]
        context.user_data["has_more_remote"] = bool(chunk["has_more"])

    results = context.user_data.get("results", [])
    index = context.user_data.get("index", 0)

    blocks = build_result_blocks(results, index, DISPLAY_PAGE_SIZE)
    if not blocks:
        await context.bot.send_message(
            chat_id=chat_id,
            text="Plus de résultats disponibles.",
            parse_mode=ParseMode.HTML,
        )
        return

    full_text = "\n\n".join(blocks)
    chunks = split_long_message(full_text)

    new_index = min(index + DISPLAY_PAGE_SIZE, len(results))
    context.user_data["index"] = new_index

    has_local_more = new_index < len(results)
    has_remote_more = context.user_data.get("has_more_remote", False)
    show_button = has_local_more or has_remote_more
    markup = pagination_keyboard() if show_button else None

    for i, chunk_text in enumerate(chunks):
        await context.bot.send_message(
            chat_id=chat_id,
            text=chunk_text,
            reply_markup=markup if i == len(chunks) - 1 else None,
            disable_web_page_preview=True,
            parse_mode=ParseMode.HTML,
        )

    if context.user_data.get("index", 0) >= len(context.user_data.get("results", [])):
        start_prefetch(context)


async def render_company_page(message_obj, context: ContextTypes.DEFAULT_TYPE) -> None:
    companies = context.user_data.get("company_results", [])
    current_page = safe_int(context.user_data.get("company_page", 0))

    if not companies:
        text = "Plus de résultats disponibles."
        markup = None
    else:
        total_pages = (len(companies) + COMPANY_PAGE_SIZE - 1) // COMPANY_PAGE_SIZE
        current_page = max(0, min(current_page, max(total_pages - 1, 0)))
        context.user_data["company_page"] = current_page

        text = build_company_page_text(companies, current_page)
        markup = company_page_keyboard(current_page, len(companies))

    await safe_edit_message_text(
        message_obj=message_obj,
        text=text,
        reply_markup=markup,
        disable_web_page_preview=True,
        parse_mode=ParseMode.HTML,
    )

# --------------------------------------------------
# ÉTAT
# --------------------------------------------------
def reset_user_flow(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data["flow"] = "idle"
    context.user_data["search_mode"] = None
    context.user_data["base_value"] = None
    context.user_data["keyword"] = None
    context.user_data["person_name"] = None
    context.user_data["fuzzy_enabled"] = False
    context.user_data["filters"] = {}
    context.user_data["results"] = []
    context.user_data["index"] = 0
    context.user_data["next_start"] = 0
    context.user_data["has_more_remote"] = True
    context.user_data["prefetched_chunk"] = None
    context.user_data["prefetch_task"] = None
    context.user_data["page_lock"] = asyncio.Lock()
    context.user_data["company_results"] = []
    context.user_data["company_page"] = 0
    context.user_data["company_list_message_id"] = None
    context.user_data.setdefault("last_search_summary", None)
    context.user_data.setdefault("last_search_export_query", None)
    context.user_data.setdefault("last_search_mode", None)
    context.user_data.setdefault("last_search", None)
    context.user_data["max_results"] = 50
    context.user_data["prospect_mode_variant"] = "quick"

# --------------------------------------------------
# ENVOI BRANDING
# --------------------------------------------------
async def send_brand_welcome(message_obj) -> None:
    caption = f"{brand_header_text()}\n\n{main_menu_text()}"

    if BRAND_IMAGE:
        try:
            if os.path.exists(BRAND_IMAGE):
                with open(BRAND_IMAGE, "rb") as img:
                    await message_obj.reply_photo(
                        photo=img,
                        caption=caption,
                        parse_mode=ParseMode.HTML,
                        reply_markup=menu_keyboard(),
                    )
            else:
                await message_obj.reply_photo(
                    photo=BRAND_IMAGE,
                    caption=caption,
                    parse_mode=ParseMode.HTML,
                    reply_markup=menu_keyboard(),
                )
            return
        except Exception as e:
            logger.warning("Impossible d'envoyer la bannière: %s", e)

    await message_obj.reply_text(
        caption,
        parse_mode=ParseMode.HTML,
        reply_markup=menu_keyboard(),
        disable_web_page_preview=True,
    )

# --------------------------------------------------
# COMMANDES
# --------------------------------------------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_access(update, context):
        return
    reset_user_flow(context)
    if update.message:
        await send_brand_welcome(update.message)


async def prospects(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_access(update, context):
        return
    reset_user_flow(context)
    await update.message.reply_text(
        prospects_intro_text(),
        parse_mode=ParseMode.HTML,
        reply_markup=prospects_mode_keyboard(),
    )


async def person_search(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_access(update, context):
        return
    reset_user_flow(context)
    await update.message.reply_text(
        person_intro_text(),
        parse_mode=ParseMode.HTML,
        reply_markup=person_menu_keyboard(),
    )


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_access(update, context):
        return
    reset_user_flow(context)
    await update.message.reply_text(
        "❌ <b>Recherche annulée.</b>\nTu peux relancer avec /start.",
        parse_mode=ParseMode.HTML,
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_access(update, context):
        return
    await update.message.reply_text(
        "<b>Commandes disponibles</b>\n\n"
        "/start\n"
        "/help\n"
        "/prospects\n"
        "/personne\n"
        "/cancel\n\n"
        "<b>Modules</b>\n"
        "• Recherche prospects LinkedIn\n"
        "• Recherche personne sur LinkedIn\n"
        "• Recherche entreprise\n\n"
        "<b>Exemples recherche entreprise</b>\n"
        "<code>Dupont</code>\n"
        "<code>Dupont Martin</code>\n"
        "<code>Dupont Martin,ville=Paris</code>",
        parse_mode=ParseMode.HTML,
    )

# --------------------------------------------------
# TEXTE
# --------------------------------------------------
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message:
        return
    if not await require_access(update, context):
        return

    flow = context.user_data.get("flow", "idle")
    text = (update.message.text or "").strip()

    if flow in {"awaiting_keyword", "awaiting_keyword_quick", "awaiting_keyword_advanced"}:
        if not text:
            await update.message.reply_text(
                "Merci d'envoyer un mot-clé valide.",
                parse_mode=ParseMode.HTML,
            )
            return

        context.user_data["keyword"] = text
        context.user_data["base_value"] = text
        mode_variant = context.user_data.get("prospect_mode_variant", "quick")
        if mode_variant == "advanced" or flow == "awaiting_keyword_advanced":
            context.user_data["flow"] = "awaiting_filters"
            await update.message.reply_text(
                filters_help_text(),
                parse_mode=ParseMode.HTML,
                reply_markup=back_to_menu_keyboard(),
            )
        else:
            context.user_data["filters"] = {}
            context.user_data["flow"] = "awaiting_prospect_geo"
            await update.message.reply_text(
                "<b>Zone géographique</b>\n\nChoisis une préférence pour équilibrer les résultats :",
                parse_mode=ParseMode.HTML,
                reply_markup=prospect_geo_keyboard(),
            )
        return

    if flow == "awaiting_filters":
        context.user_data["filters"] = parse_filters(text)
        context.user_data["flow"] = "awaiting_prospect_geo"

        await update.message.reply_text(
            "<b>Zone géographique</b>\n\nChoisis une préférence pour équilibrer les résultats :",
            parse_mode=ParseMode.HTML,
            reply_markup=prospect_geo_keyboard(),
        )
        return

    if flow == "awaiting_person_name":
        if not text:
            await update.message.reply_text(
                "Merci d'envoyer un nom et prénom valides.",
                parse_mode=ParseMode.HTML,
            )
            return

        context.user_data["person_name"] = text
        context.user_data["base_value"] = text
        context.user_data["flow"] = "awaiting_person_fuzzy"

        await update.message.reply_text(
            "<b>Orthographe proche</b>\n\n"
            "Souhaites-tu prendre en compte une orthographe approchante ?\n"
            "Exemple : Sarah / Sarra / Sara",
            parse_mode=ParseMode.HTML,
            reply_markup=fuzzy_keyboard(),
        )
        return

    if flow == "awaiting_person_filters":
        context.user_data["filters"] = parse_filters(text)
        context.user_data["flow"] = "awaiting_person_max_results"

        await update.message.reply_text(
            "Combien de résultats maximum veux-tu récupérer ?",
            parse_mode=ParseMode.HTML,
            reply_markup=max_results_keyboard(),
        )
        return

    if flow == "awaiting_company_query":
        name, company_filters = parse_company_query_input(text)

        if not name:
            await update.message.reply_text(
                "Merci d'envoyer un nom ou un nom prénom valide.\n"
                "Exemple : <code>Dupont,ville=Paris</code>",
                parse_mode=ParseMode.HTML,
            )
            return

        ville_filter = company_filters.get("ville", "")

        await update.message.reply_text(
            company_search_loading_text(),
            parse_mode=ParseMode.HTML,
        )

        companies = await asyncio.to_thread(search_company_person, name, ville_filter)

        if not companies:
            await update.message.reply_text(
                no_result_text(),
                parse_mode=ParseMode.HTML,
            )
            reset_user_flow(context)
            return

        context.user_data["company_results"] = companies
        context.user_data["company_page"] = 0
        context.user_data["flow"] = "company_results_ready"

        context.user_data["last_search_summary"] = search_summary_text("company", name, {k: v for k, v in company_filters.items() if v}, len(companies))
        context.user_data["last_search_export_query"] = name
        context.user_data["last_search_mode"] = "company"
        await update.message.reply_text(
            context.user_data["last_search_summary"],
            parse_mode=ParseMode.HTML,
        )
        await after_search_usage(update, context)

        page_message = await update.message.reply_text(
            "Chargement des résultats…",
            parse_mode=ParseMode.HTML,
        )
        context.user_data["company_list_message_id"] = page_message.message_id
        await render_company_page(page_message, context)
        return

    await update.message.reply_text(
        "Utilise /start, /prospects ou /personne pour démarrer.",
        parse_mode=ParseMode.HTML,
    )

# --------------------------------------------------
# CALLBACKS
# --------------------------------------------------
async def launch_search(query, context: ContextTypes.DEFAULT_TYPE, export_excel_requested: bool) -> None:
    search_mode = context.user_data.get("search_mode")
    base_value = context.user_data.get("base_value")
    custom_filters = context.user_data.get("filters", {})
    fuzzy_enabled = bool(context.user_data.get("fuzzy_enabled", False))

    if not search_mode or not base_value:
        await query.message.reply_text(
            "Recherche introuvable. Relance avec /start.",
            parse_mode=ParseMode.HTML,
        )
        reset_user_flow(context)
        return

    context.user_data["flow"] = "searching"
    await context.bot.send_chat_action(chat_id=query.message.chat_id, action=ChatAction.TYPING)

    await query.message.reply_text(
        loading_search_text(search_mode),
        parse_mode=ParseMode.HTML,
    )

    try:
        first_chunk = await fetch_display_chunk(
            search_mode=search_mode,
            base_value=base_value,
            custom_filters=custom_filters,
            start=0,
            target_count=DISPLAY_PAGE_SIZE,
            fuzzy_enabled=fuzzy_enabled,
        )
    except Exception as e:
        logger.exception("Erreur pendant la recherche initiale")
        await query.message.reply_text(
            f"Erreur pendant la recherche : <code>{esc(e)}</code>",
            parse_mode=ParseMode.HTML,
        )
        reset_user_flow(context)
        return

    first_results = first_chunk.get("results", [])
    if not first_results:
        await query.message.reply_text(
            no_result_text(),
            parse_mode=ParseMode.HTML,
        )
        reset_user_flow(context)
        return

    context.user_data["results"] = first_results[:]
    context.user_data["index"] = 0
    context.user_data["next_start"] = first_chunk["next_start"]
    context.user_data["has_more_remote"] = bool(first_chunk["has_more"])
    context.user_data["prefetched_chunk"] = None
    context.user_data["prefetch_task"] = None
    context.user_data["flow"] = "results_ready"
    context.user_data["last_search_summary"] = search_summary_text(search_mode, base_value, custom_filters, len(first_results))
    context.user_data["last_search_export_query"] = base_value
    context.user_data["last_search_mode"] = search_mode
    context.user_data["last_search"] = {"mode": search_mode, "base_value": base_value, "filters": custom_filters, "fuzzy_enabled": fuzzy_enabled, "summary": context.user_data["last_search_summary"], "count": len(first_results)}

    await query.message.reply_text(
        context.user_data["last_search_summary"],
        parse_mode=ParseMode.HTML,
    )
    fake_update = Update(update_id=0, message=query.message)
    await after_search_usage(fake_update, context)
    await send_next_page(query.message.chat_id, context)
    start_prefetch(context)

    if export_excel_requested:
        await query.message.reply_text(
            "📊 <b>Génération de l’Excel complet…</b>",
            parse_mode=ParseMode.HTML,
        )
        try:
            full_results = await asyncio.to_thread(
                search_full_export,
                search_mode,
                base_value,
                custom_filters,
                int(context.user_data.get("max_results", MAX_RESULTS)),
                fuzzy_enabled,
            )
            file_path = export_excel(full_results, base_value)
            with open(file_path, "rb") as f:
                await query.message.reply_document(
                    document=f,
                    filename=os.path.basename(file_path),
                    caption=f"Voici ton export Excel complet ({len(full_results)} profils).",
                )
        except Exception as e:
            logger.exception("Erreur export Excel complet")
            await query.message.reply_text(
                f"Impossible de générer l’Excel : <code>{esc(e)}</code>",
                parse_mode=ParseMode.HTML,
            )

async def last_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_access(update, context):
        return

    last = context.user_data.get("last_search")

    if not last:
        await update.message.reply_text("❌ Aucune recherche récente.")
        return

    text = (
        "🕘 <b>Dernière recherche</b>\n\n"
        f"🔎 Mot-clé : {last.get('keyword', 'N/A')}\n"
        f"📍 Zone : {last.get('zone', 'N/A')}\n"
        f"📊 Résultats : {len(last.get('results', []))}\n\n"
        "👉 Utilise les boutons pour relancer ou modifier."
    )

    await update.message.reply_text(text, parse_mode=ParseMode.HTML, reply_markup=last_search_keyboard(bool(last.get("mode") and last.get("base_value"))))

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not query:
        return

    await query.answer()
    data = query.data or ""
    flow = context.user_data.get("flow", "idle")

    if data == "access_request":
        requester = update.effective_user
        if not requester:
            return
        if has_runtime_access(requester.id):
            await query.message.reply_text("✅ Tu as déjà accès au bot. Utilise /start.", parse_mode=ParseMode.HTML)
            return

        user_info = build_user_info(requester)
        created = register_pending_request(user_info)
        sent = await send_access_request_to_owner(context.bot, user_info)

        if created and sent:
            await query.message.reply_text(
                "⏳ Demande envoyée. Je viens de prévenir l'administrateur en privé.",
                parse_mode=ParseMode.HTML,
            )
        elif get_pending_record(requester.id):
            await query.message.reply_text(
                "⏳ Ta demande est déjà en attente.",
                parse_mode=ParseMode.HTML,
            )
        else:
            await query.message.reply_text(
                "⚠️ Impossible d'envoyer la demande automatiquement pour le moment.",
                parse_mode=ParseMode.HTML,
            )
        return

    if data.startswith("access:"):
        if not await require_admin(update):
            return

        parts = data.split(":")
        if len(parts) < 3:
            await query.message.reply_text("Action d'accès invalide.", parse_mode=ParseMode.HTML)
            return

        action = parts[1]
        mode = parts[2] if len(parts) > 2 else ""
        try:
            target_user_id = int(parts[3]) if len(parts) > 3 else int(parts[2])
        except Exception:
            await query.message.reply_text("Utilisateur introuvable.", parse_mode=ParseMode.HTML)
            return

        target_info = get_pending_record(target_user_id) or get_approved_record(target_user_id) or {
            "user_id": target_user_id,
            "first_name": "Utilisateur",
            "username": "",
        }

        if action == "grant":
            record = grant_user_access(target_info, update.effective_user.id, mode)
            await notify_user_access_granted(context.bot, target_user_id, record)
            await query.message.reply_text(
                (
                    "✅ <b>Accès accordé</b>\n\n"
                    f"Utilisateur : {esc(user_display_name(target_info.get('first_name', ''), target_info.get('username', '')))}\n"
                    f"ID : <code>{target_user_id}</code>\n"
                    f"Statut : {format_access_badge(record)}"
                ),
                parse_mode=ParseMode.HTML,
                reply_markup=access_manage_keyboard(target_user_id, pending=False),
            )
            return

        if action == "deny":
            deny_user_request(target_user_id, denied_by=update.effective_user.id)
            await notify_user_access_denied(context.bot, target_user_id)
            await query.message.reply_text(
                f"❌ Demande refusée pour <code>{target_user_id}</code>.",
                parse_mode=ParseMode.HTML,
            )
            return

        if action == "revoke":
            revoke_user_access(target_user_id, revoked_by=update.effective_user.id)
            await notify_user_access_revoked(context.bot, target_user_id)
            await query.message.reply_text(
                f"🗑️ Accès supprimé pour <code>{target_user_id}</code>.",
                parse_mode=ParseMode.HTML,
            )
            return

        if action == "blacklist":
            record = blacklist_user_access(target_info, blacklisted_by=update.effective_user.id)
            await notify_user_access_revoked(context.bot, target_user_id)
            await query.message.reply_text(
                (
                    "🚫 <b>Utilisateur blacklisté</b>\n\n"
                    f"Utilisateur : {esc(user_display_name(target_info.get('first_name', ''), target_info.get('username', '')))}\n"
                    f"ID : <code>{target_user_id}</code>"
                ),
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup(profile_action_rows(record) + access_manage_keyboard(target_user_id, pending=False, blacklisted=True).inline_keyboard),
            )
            return

        if action == "unblacklist":
            unblacklist_user_access(target_user_id, actor_id=update.effective_user.id)
            await query.message.reply_text(
                f"♻️ Blacklist retirée pour <code>{target_user_id}</code>.",
                parse_mode=ParseMode.HTML,
            )
            return

    if data == "menu_home":
        reset_user_flow(context)
        await query.message.reply_text(main_menu_text(), parse_mode=ParseMode.HTML, reply_markup=menu_keyboard())
        return

    if data == "menu_help":
        if not await require_access(update, context):
            return
        await query.message.reply_text(
            f"<b>{BOT_BRAND_NAME}</b>\n\n"
            "Choisis un module puis suis les étapes.\n"
            "Le mode prospects avancé te permet aussi de choisir France, Maghreb, Europe ou aucun filtre.",
            parse_mode=ParseMode.HTML,
            reply_markup=back_to_menu_keyboard(),
        )
        return

    if data == "menu_last":
        if not await require_access(update, context):
            return
        summary = context.user_data.get("last_search_summary") or "Aucune recherche récente disponible."
        await query.message.reply_text(summary, parse_mode=ParseMode.HTML, reply_markup=last_search_keyboard(bool(context.user_data.get("last_search_mode") and context.user_data.get("last_search_export_query"))))
        return

    if data == "action_rerun_last":
        if not await require_access(update, context):
            return
        last = context.user_data.get("last_search") or {}
        search_mode = last.get("mode") or context.user_data.get("last_search_mode")
        base_value = last.get("base_value") or context.user_data.get("last_search_export_query")
        custom_filters = last.get("filters") or context.user_data.get("filters", {})
        fuzzy_enabled = bool(last.get("fuzzy_enabled", context.user_data.get("fuzzy_enabled", False)))
        if not search_mode or not base_value:
            await query.message.reply_text("Aucune recherche relançable disponible.", parse_mode=ParseMode.HTML)
            return
        context.user_data["search_mode"] = search_mode
        context.user_data["base_value"] = base_value
        context.user_data["filters"] = custom_filters
        context.user_data["fuzzy_enabled"] = fuzzy_enabled
        if search_mode == "company":
            context.user_data["flow"] = "awaiting_company_query"
            fake_text = Update(update_id=0, callback_query=query)
            companies = await asyncio.to_thread(search_company_person, base_value, custom_filters.get("ville", ""))
            context.user_data["company_results"] = companies
            context.user_data["company_page"] = 0
            context.user_data["flow"] = "company_results_ready"
            await query.message.reply_text(context.user_data.get("last_search_summary") or search_summary_text("company", base_value, custom_filters, len(companies)), parse_mode=ParseMode.HTML)
            await render_company_page(query.message, context)
            return
        context.user_data["flow"] = "results_ready"
        await launch_search(query, context, False)
        return

    if data == "action_export_last":
        if not await require_access(update, context):
            return
        search_mode = context.user_data.get("last_search_mode")
        base_value = context.user_data.get("last_search_export_query")
        custom_filters = context.user_data.get("filters", {})
        fuzzy_enabled = bool(context.user_data.get("fuzzy_enabled", False))
        if not search_mode or not base_value:
            await query.message.reply_text("Aucune recherche exportable disponible.", parse_mode=ParseMode.HTML)
            return
        await query.message.reply_text("📊 <b>Génération de l’Excel complet…</b>", parse_mode=ParseMode.HTML)
        try:
            full_results = await asyncio.to_thread(search_full_export, search_mode, base_value, custom_filters, int(context.user_data.get("max_results", MAX_RESULTS)), fuzzy_enabled)
            file_path = export_excel(full_results, base_value)
            with open(file_path, "rb") as f:
                await query.message.reply_document(document=f, filename=os.path.basename(file_path), caption=f"Voici ton export Excel complet ({len(full_results)} profils).")
        except Exception as e:
            await query.message.reply_text(f"Impossible de générer l’Excel : <code>{esc(e)}</code>", parse_mode=ParseMode.HTML)
        return

    if not await require_access(update, context):
        return

    if data == "menu_prospects":
        reset_user_flow(context)
        await query.message.reply_text(prospects_intro_text(), parse_mode=ParseMode.HTML, reply_markup=prospects_mode_keyboard())
        return

    if data == "prospect_mode_quick":
        reset_user_flow(context)
        context.user_data["search_mode"] = "prospect"
        context.user_data["prospect_mode_variant"] = "quick"
        context.user_data["flow"] = "awaiting_keyword_quick"
        await query.message.reply_text(prospect_query_help_text("quick"), parse_mode=ParseMode.HTML, reply_markup=last_search_keyboard(bool(context.user_data.get("last_search_mode") and context.user_data.get("last_search_export_query"))))
        return

    if data == "prospect_mode_advanced":
        reset_user_flow(context)
        context.user_data["search_mode"] = "prospect"
        context.user_data["prospect_mode_variant"] = "advanced"
        context.user_data["flow"] = "awaiting_keyword_advanced"
        await query.message.reply_text(prospect_query_help_text("advanced"), parse_mode=ParseMode.HTML, reply_markup=last_search_keyboard(bool(context.user_data.get("last_search_mode") and context.user_data.get("last_search_export_query"))))
        return

    if data == "menu_company_direct":
        reset_user_flow(context)
        context.user_data["flow"] = "awaiting_company_query"
        context.user_data["search_mode"] = "company"
        await query.message.reply_text(
            "<b>Recherche entreprise</b>\n\n"
            "Entre un nom ou un nom prénom.\n"
            "Tu peux ajouter une ville si tu veux.\n\n"
            "Exemple : <code>Dupont,ville=Paris</code>",

            parse_mode=ParseMode.HTML,
            reply_markup=back_to_menu_keyboard(),
        )
        return

    if data == "menu_person_menu":
        reset_user_flow(context)
        await query.message.reply_text(
            person_intro_text(),
            parse_mode=ParseMode.HTML,
            reply_markup=person_menu_keyboard(),
        )
        return

    if data == "person_linkedin":
        reset_user_flow(context)
        context.user_data["flow"] = "awaiting_person_name"
        context.user_data["search_mode"] = "person"
        await query.message.reply_text(
            "<b>Recherche LinkedIn</b>\n\nQuel nom et prénom veux-tu rechercher ?",
            parse_mode=ParseMode.HTML,
        )
        return

    if data == "person_company":
        reset_user_flow(context)
        context.user_data["flow"] = "awaiting_company_query"
        context.user_data["search_mode"] = "company"
        await query.message.reply_text(
            "<b>Recherche entreprise</b>\n\n"
            "Entre un nom ou un nom prénom.\n"
            "Tu peux ajouter une ville si tu veux.\n\n"
            "Exemple : <code>Dupont,ville=Paris</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    if data == "more":
        lock = context.user_data.get("page_lock")
        if lock is None:
            context.user_data["page_lock"] = asyncio.Lock()
            lock = context.user_data["page_lock"]

        async with lock:
            await send_next_page(query.message.chat_id, context)
        return

    if data == "company_next":
        if context.user_data.get("flow") != "company_results_ready":
            await query.message.reply_text(
                "Cette action n'est plus valide. Relance avec /personne.",
                parse_mode=ParseMode.HTML,
            )
            return

        companies = context.user_data.get("company_results", [])
        total_pages = max(1, (len(companies) + COMPANY_PAGE_SIZE - 1) // COMPANY_PAGE_SIZE)
        current_page = safe_int(context.user_data.get("company_page", 0))

        if current_page >= total_pages - 1:
            return

        context.user_data["company_page"] = current_page + 1
        await render_company_page(query.message, context)
        return

    if data == "company_prev":
        if context.user_data.get("flow") != "company_results_ready":
            await query.message.reply_text(
                "Cette action n'est plus valide. Relance avec /personne.",
                parse_mode=ParseMode.HTML,
            )
            return

        current_page = safe_int(context.user_data.get("company_page", 0))
        if current_page <= 0:
            return

        context.user_data["company_page"] = current_page - 1
        await render_company_page(query.message, context)
        return

    if data == "company_back":
        if context.user_data.get("flow") != "company_results_ready":
            await query.message.reply_text(
                "Cette action n'est plus valide. Relance avec /personne.",
                parse_mode=ParseMode.HTML,
            )
            return

        try:
            await query.message.delete()
        except BadRequest:
            pass
        return

    if data.startswith("company_detail_"):
        try:
            idx = int(data.split("_")[-1])
        except Exception:
            await query.message.reply_text(
                "Impossible d’ouvrir le détail de ce résultat.",
                parse_mode=ParseMode.HTML,
            )
            return

        company_results = context.user_data.get("company_results", [])
        if idx < 0 or idx >= len(company_results):
            await query.message.reply_text(
                "Ce résultat n’est plus disponible. Relance la recherche.",
                parse_mode=ParseMode.HTML,
            )
            return

        row = company_results[idx]
        detail_text = build_company_detail_text(idx + 1, row)

        await query.message.reply_text(
            detail_text,
            reply_markup=company_detail_keyboard(),
            disable_web_page_preview=True,
            parse_mode=ParseMode.HTML,
        )
        return

    if data.startswith("prospect_geo_"):
        if flow != "awaiting_prospect_geo":
            await query.message.reply_text("Cette action n'est plus valide. Relance avec /prospects.", parse_mode=ParseMode.HTML)
            return
        preset = data.replace("prospect_geo_", "")
        context.user_data["filters"] = apply_prospect_geo_filter(context.user_data.get("filters", {}), preset)
        context.user_data["flow"] = "awaiting_max_results"
        await query.message.reply_text(
            f"<b>Zone sélectionnée</b> : {esc(prospect_geo_label(preset))}\n\n"
            "Combien de résultats maximum veux-tu récupérer ?",
            parse_mode=ParseMode.HTML,
            reply_markup=max_results_keyboard(),
        )
        return

    if data.startswith("max_results:"):
        try:
            value = int(data.split(":", 1)[1])
        except Exception:
            value = 50
        context.user_data["max_results"] = value
        if flow == "awaiting_max_results":
            context.user_data["flow"] = "awaiting_excel_choice"
        elif flow == "awaiting_person_max_results":
            context.user_data["flow"] = "awaiting_person_excel_choice"
        else:
            await query.message.reply_text("Cette action n'est plus valide. Relance avec /start.", parse_mode=ParseMode.HTML)
            return
        await query.message.reply_text(excel_choice_text(), parse_mode=ParseMode.HTML, reply_markup=excel_keyboard())
        return

    if data in {"fuzzy_yes", "fuzzy_no"}:
        if flow != "awaiting_person_fuzzy":
            await query.message.reply_text(
                "Cette action n'est plus valide. Relance avec /personne.",
                parse_mode=ParseMode.HTML,
            )
            return

        context.user_data["fuzzy_enabled"] = (data == "fuzzy_yes")
        context.user_data["flow"] = "awaiting_person_filters"

        await query.message.reply_text(
            person_filters_help_text(),
            parse_mode=ParseMode.HTML,
            reply_markup=back_to_menu_keyboard(),
        )
        return

    if data in {"excel_yes", "excel_no"}:
        export_excel_requested = data == "excel_yes"

        if flow == "awaiting_excel_choice":
            await launch_search(query, context, export_excel_requested)
            return

        if flow == "awaiting_person_excel_choice":
            await launch_search(query, context, export_excel_requested)
            return

        await query.message.reply_text(
            "Cette action n'est plus valide. Relance avec /start.",
            parse_mode=ParseMode.HTML,
        )
        return

# --------------------------------------------------
# ERREURS
# --------------------------------------------------
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.exception("Erreur non gérée", exc_info=context.error)
    try:
        if isinstance(update, Update) and update.effective_chat:
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="⚠️ Une erreur inattendue s'est produite. Réessaie avec /start.",
                parse_mode=ParseMode.HTML,
            )
    except Exception:
        pass

# --------------------------------------------------
# MAIN
# --------------------------------------------------
def main() -> None:
    load_cache()
    load_access_state()

    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("prospects", prospects))
    app.add_handler(CommandHandler("personne", person_search))
    app.add_handler(CommandHandler("cancel", cancel))
    app.add_handler(CommandHandler("last", last_command))
    app.add_handler(CommandHandler("admin", admin_command))
    app.add_handler(CommandHandler("stats", stats_command))
    app.add_handler(CommandHandler("whoami", whoami_command))
    app.add_handler(CommandHandler("cache_clear", cache_clear_command))
    app.add_handler(CommandHandler("users", users_command))
    app.add_handler(CommandHandler("pending", pending_command))
    app.add_handler(CommandHandler("blacklist", blacklist_command))
    app.add_handler(CommandHandler("unblacklist", unblacklist_command))
    app.add_handler(CommandHandler("blacklisted", blacklist_list_command))
    app.add_handler(CommandHandler("history", history_command))

    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    app.add_error_handler(error_handler)

    print(f"{BOT_BRAND_NAME} actif")
    app.run_polling()


if __name__ == "__main__":
    main()

